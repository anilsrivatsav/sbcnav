from __future__ import annotations

import argparse
import io
import json
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func

from database import SessionLocal, engine
from models import CateringSyncRun, DataChangeLog, Earning, EarningLink, Station, Unit
from services import hash_row

UNIT_SHEET = "UNITS BASE DATA"
EARNING_SHEET = "EARNINGS BASE DATA"

UNIT_FIELDS = (
    "sl_no",
    "unit_no",
    "type_of_unit",
    "station_code",
    "station_category",
    "old_category",
    "pf_no",
    "pegged_location",
    "reservation_category",
    "allotment_type",
    "licensee_name",
    "license_fee",
    "contract_from",
    "contract_to",
    "paid_upto",
    "unit_status",
)

EARNING_FIELDS = (
    "sl_no",
    "date_of_receipt",
    "raw_unit_no",
    "raw_station_code",
    "pf_no",
    "licensee_name",
    "payment_head",
    "payment_sub_head",
    "period_from",
    "period_to",
    "amount",
    "gst",
    "receipt_type",
    "mr_no",
    "mr_date",
    "ua_case",
)

EXPECTED_UNIT_HEADERS = {
    1: "SL NO.",
    2: "UNIT NO.",
    3: "TYPE OF UNIT",
    4: "STATION",
    5: "STATION CATEGORY",
    7: "PF NO",
    8: "PEGGED LOCATION",
    9: "RESERVATION CATEGORY",
    10: "TYPE OF ALLOTMENT",
    11: "NAME OF LICENSEE",
    12: "LICENSE FEE",
    13: "CONTRACT PERIOD",
    15: "LICENSE FEE",
    16: "UNIT STATUS",
}

EXPECTED_EARNING_HEADERS = {
    1: "SL. NO.",
    2: "DATE OF RECEIPT",
    3: "UNIT NO.",
    4: "STATION",
    5: "PF NO.",
    6: "NAME OF LICENSEE",
    7: "PAYMENT HEAD",
    8: "PAYMENT SUB-HEAD",
    9: "PERIOD",
    11: "AMOUNT",
    12: "GST",
    13: "RECIEPT TYPE",
    14: "MR NO/UTS NO/ CHALLAN NO",
    15: "MR DATE",
    16: "U/A CASE",
}


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def identifier(value: Any) -> str | None:
    text = clean(value)
    return text.upper() if text else None


def integer(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).replace(",", "").replace("INR", "").replace("Rs.", "").strip()
    try:
        return int(round(float(text)))
    except ValueError:
        return None


def date_iso(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    if not text:
        return None
    for pattern in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return text


def normalized_header(value: Any) -> str:
    return " ".join((clean(value) or "").upper().split())


def validate_headers(sheet, expected: dict[int, str], subheaders: dict[int, str]) -> None:
    errors: list[str] = []
    for column, wanted in expected.items():
        actual = normalized_header(sheet.cell(2, column).value)
        if actual != wanted:
            errors.append(f"{sheet.title}!{sheet.cell(2, column).coordinate}: expected {wanted!r}, got {actual!r}")
    for column, wanted in subheaders.items():
        actual = normalized_header(sheet.cell(3, column).value)
        if actual != wanted:
            errors.append(f"{sheet.title}!{sheet.cell(3, column).coordinate}: expected {wanted!r}, got {actual!r}")
    if errors:
        raise ValueError("Workbook layout validation failed:\n" + "\n".join(errors))


def optional_header_column(sheet, *names: str) -> int | None:
    wanted = {normalized_header(name) for name in names}
    for row in (2, 3):
        for column in range(1, sheet.max_column + 1):
            if normalized_header(sheet.cell(row, column).value) in wanted:
                return column
    return None


def parse_units(sheet) -> tuple[list[dict[str, Any]], dict[str, int]]:
    rows: list[dict[str, Any]] = []
    skipped_labels = 0
    seen: set[str] = set()
    remarks_column = optional_header_column(sheet, "REMARKS", "REMARK")
    for row_number in range(4, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, 17)]
        if not any(value not in (None, "") for value in values):
            continue
        raw = dict(zip(UNIT_FIELDS, values, strict=True))
        unit_no = identifier(raw["unit_no"])
        if not unit_no:
            skipped_labels += 1
            continue
        if unit_no in seen:
            raise ValueError(f"Duplicate unit number {unit_no!r} at {UNIT_SHEET} row {row_number}")
        seen.add(unit_no)
        raw_status = clean(raw["unit_status"])
        contract_from = date_iso(raw["contract_from"])
        contract_to = date_iso(raw["contract_to"])
        licensee_name = clean(raw["licensee_name"])
        available = not licensee_name and not contract_from and not contract_to
        explicit_remarks = (
            clean(sheet.cell(row_number, remarks_column).value)
            if remarks_column is not None
            else None
        )
        rows.append(
            {
                "sl_no": integer(raw["sl_no"]),
                "unit_no": unit_no,
                "type_of_unit": clean(raw["type_of_unit"]),
                "station_code": identifier(raw["station_code"]),
                "station_category": clean(raw["station_category"]),
                "old_category": clean(raw["old_category"]),
                "pf_no": clean(raw["pf_no"]),
                "pegged_location": clean(raw["pegged_location"]),
                "reservation_category": clean(raw["reservation_category"]),
                "allotment_type": clean(raw["allotment_type"]),
                "licensee_name": licensee_name,
                "license_fee": clean(integer(raw["license_fee"])),
                "contract_from": contract_from,
                "contract_to": contract_to,
                "paid_upto": date_iso(raw["paid_upto"]),
                "unit_status": "Available" if available else raw_status,
                "remarks": explicit_remarks or (raw_status if available else None),
                "source_row": row_number,
            }
        )
    return rows, {"skipped_category_labels": skipped_labels}


def parse_earnings(sheet) -> list[dict[str, Any]]:
    rows_by_fingerprint: dict[str, dict[str, Any]] = {}
    seen_serials: set[int] = set()
    for row_number in range(4, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, 17)]
        if not any(value not in (None, "") for value in values):
            continue
        raw = dict(zip(EARNING_FIELDS, values, strict=True))
        sl_no = integer(raw["sl_no"])
        if sl_no is None:
            raise ValueError(f"Missing earning serial number at {EARNING_SHEET} row {row_number}")
        if sl_no in seen_serials:
            raise ValueError(f"Duplicate earning serial number {sl_no} at {EARNING_SHEET} row {row_number}")
        seen_serials.add(sl_no)
        source = {
            "sl_no": sl_no,
            "date_of_receipt": date_iso(raw["date_of_receipt"]),
            "raw_unit_no": identifier(raw["raw_unit_no"]),
            "raw_station_code": identifier(raw["raw_station_code"]),
            "pf_no": clean(raw["pf_no"]),
            "licensee_name": clean(raw["licensee_name"]),
            "payment_head": clean(raw["payment_head"]),
            "payment_sub_head": clean(raw["payment_sub_head"]),
            "period_from": date_iso(raw["period_from"]),
            "period_to": date_iso(raw["period_to"]),
            "amount": integer(raw["amount"]),
            "gst": integer(raw["gst"]),
            "receipt_type": clean(raw["receipt_type"]),
            "mr_no": clean(raw["mr_no"]),
            "mr_date": date_iso(raw["mr_date"]),
            "ua_case": clean(raw["ua_case"]),
        }
        fingerprint_payload = {key: value for key, value in source.items() if key != "sl_no"}
        receipt_key = hash_row("earning-base-data-v2", fingerprint_payload)
        existing = rows_by_fingerprint.get(receipt_key)
        if existing:
            existing["_source_rows"].append(row_number)
            existing["duplicate_count"] += 1
            continue
        source["receipt_key"] = receipt_key
        source["duplicate_count"] = 1
        source["_source_rows"] = [row_number]
        rows_by_fingerprint[receipt_key] = source

    rows = []
    for source in rows_by_fingerprint.values():
        source["source_rows"] = ",".join(str(value) for value in source.pop("_source_rows"))
        rows.append(source)
    return rows


def load_source(source: Path | bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    workbook_source = io.BytesIO(source) if isinstance(source, bytes) else source
    workbook = load_workbook(workbook_source, data_only=True)
    missing_sheets = [name for name in (UNIT_SHEET, EARNING_SHEET) if name not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"Missing required worksheet(s): {', '.join(missing_sheets)}")

    unit_sheet = workbook[UNIT_SHEET]
    earning_sheet = workbook[EARNING_SHEET]
    validate_headers(unit_sheet, EXPECTED_UNIT_HEADERS, {13: "FROM", 14: "TO", 15: "PAID UPTO"})
    validate_headers(earning_sheet, EXPECTED_EARNING_HEADERS, {9: "FROM", 10: "TO"})

    units, unit_notes = parse_units(unit_sheet)
    earnings = parse_earnings(earning_sheet)
    if not units:
        raise ValueError("No unit records were found")
    if not earnings:
        raise ValueError("No earning records were found")
    duplicate_rows = sum(max(row.get("duplicate_count", 1) - 1, 0) for row in earnings)
    return units, earnings, {
        **unit_notes,
        "earning_source_rows": len(earnings) + duplicate_rows,
        "duplicate_earning_rows": duplicate_rows,
        "duplicate_earning_groups": sum(1 for row in earnings if row.get("duplicate_count", 1) > 1),
    }


def reconcile(
    units: list[dict[str, Any]],
    earnings: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    with SessionLocal() as session:
        station_codes = {code for (code,) in session.query(Station.station_code).all()}
        existing_units = {unit_no for (unit_no,) in session.query(Unit.unit_no).all()}
        current = {
            "units": session.query(func.count(Unit.unit_no)).scalar() or 0,
            "earnings": session.query(func.count(Earning.earning_key)).scalar() or 0,
            "earning_links": session.query(func.count(EarningLink.id)).scalar() or 0,
        }

    source_unit_codes = {row["unit_no"] for row in units}
    source_units = {row["unit_no"]: row for row in units}
    available_unit_numbers = {
        row["unit_no"] for row in units if row.get("unit_status") == "Available"
    }
    missing_unit_stations = sorted(
        {row["station_code"] for row in units if row["station_code"] and row["station_code"] not in station_codes}
    )
    if missing_unit_stations:
        raise ValueError("Unit station codes missing from station master: " + ", ".join(missing_unit_stations))

    prepared: list[dict[str, Any]] = []
    unlinked_units: dict[str, int] = {}
    miscellaneous_labels: dict[str, int] = {}
    invalid_stations: dict[str, int] = {}
    station_mismatches: list[dict[str, Any]] = []
    for source in earnings:
        raw_unit = source["raw_unit_no"]
        raw_station = source["raw_station_code"]
        matched_unit = raw_unit if raw_unit in source_unit_codes else None
        matched_station = raw_station if raw_station in station_codes else None
        unit_station = source_units.get(matched_unit, {}).get("station_code") if matched_unit else None
        earning_scope = (
            "tender_emd"
            if matched_unit in available_unit_numbers
            else "unit"
            if matched_unit
            else "miscellaneous"
            if raw_unit in {"VENDOR'S CHECK", "WVM", "IRCTC"}
            else "unlinked"
        )
        if raw_unit and earning_scope == "unlinked":
            unlinked_units[raw_unit] = unlinked_units.get(raw_unit, 0) + 1
        elif raw_unit and earning_scope == "miscellaneous":
            miscellaneous_labels[raw_unit] = miscellaneous_labels.get(raw_unit, 0) + 1
        if raw_station and not matched_station:
            invalid_stations[raw_station] = invalid_stations.get(raw_station, 0) + 1
        if matched_unit and matched_station and unit_station and matched_station != unit_station:
            station_mismatches.append({
                "source_rows": source.get("source_rows"),
                "unit_no": matched_unit,
                "unit_station": unit_station,
                "earning_station": matched_station,
            })
        prepared.append(
            {
                "receipt_key": source["receipt_key"],
                "sl_no": source["sl_no"],
                "source_rows": source["source_rows"],
                "duplicate_count": source["duplicate_count"],
                "date_of_receipt": source["date_of_receipt"],
                "unit_no": matched_unit,
                "station_code": unit_station or matched_station,
                "raw_unit_no": raw_unit,
                "raw_station_code": raw_station,
                "earning_scope": earning_scope,
                "pf_no": source["pf_no"],
                "licensee_name": source["licensee_name"],
                "payment_head": source["payment_head"],
                "payment_sub_head": source["payment_sub_head"],
                "period_from": source["period_from"],
                "period_to": source["period_to"],
                "amount": source["amount"],
                "gst": source["gst"],
                "receipt_type": source["receipt_type"],
                "mr_no": source["mr_no"],
                "mr_date": source["mr_date"],
                "ua_case": source["ua_case"],
            }
        )

    report = {
        "database": {
            "dialect": engine.dialect.name,
            "current": current,
        },
        "source": {
            "units": len(units),
            "earnings": len(earnings),
            "earning_amount_total": sum(row["amount"] or 0 for row in prepared),
            "gst_total": sum(row["gst"] or 0 for row in prepared),
        },
        "reconciliation": {
            "new_unit_numbers": len(source_unit_codes - existing_units),
            "existing_unit_numbers": len(source_unit_codes & existing_units),
            "units_absent_from_source": len(existing_units - source_unit_codes),
            "linked_earning_rows": sum(1 for row in prepared if row["unit_no"]),
            "unlinked_earning_rows": sum(1 for row in prepared if not row["unit_no"]),
            "miscellaneous_earning_rows": sum(1 for row in prepared if row["earning_scope"] == "miscellaneous"),
            "tender_emd_rows": sum(1 for row in prepared if row["earning_scope"] == "tender_emd"),
            "available_units": len(available_unit_numbers),
            "miscellaneous_labels": dict(sorted(miscellaneous_labels.items())),
            "unlinked_unit_labels": dict(sorted(unlinked_units.items())),
            "invalid_station_labels": dict(sorted(invalid_stations.items())),
            "station_mismatch_count": len(station_mismatches),
            "station_mismatches": station_mismatches[:100],
        },
    }
    return prepared, report


def public_unit_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if not key.startswith("_")}


def apply_import(
    source_name: str,
    units: list[dict[str, Any]],
    earnings: list[dict[str, Any]],
    report: dict[str, Any],
    spreadsheet_id: str = "local-workbook",
) -> None:
    now = datetime.now(timezone.utc)
    source_unit_numbers = {row["unit_no"] for row in units}
    session = SessionLocal()
    try:
        with session.begin():
            # Links are rebuilt from the authoritative source after unit and receipt
            # reconciliation. Receipt rows themselves are updated incrementally.
            session.query(EarningLink).delete(synchronize_session=False)

            existing_units = {unit.unit_no: unit for unit in session.query(Unit).all()}
            existing_units_casefold = {unit_no.upper(): unit for unit_no, unit in existing_units.items()}
            mutable_unit_fields = [
                column.name
                for column in Unit.__table__.columns
                if column.name
                not in {"unit_no", "created_at", "first_seen_at", "updated_at", "last_seen_at", "is_active", "source_hash"}
            ]
            inserted_units = 0
            updated_units = 0
            unchanged_units = 0
            renamed_units = 0
            for source in units:
                payload = public_unit_row(source)
                payload["source_hash"] = hash_row("unit-base-data", payload)
                unit = existing_units.get(payload["unit_no"]) or existing_units_casefold.get(payload["unit_no"].upper())
                if unit is None:
                    unit = Unit(
                        **payload,
                        created_at=now,
                        updated_at=now,
                        first_seen_at=now,
                        last_seen_at=now,
                        is_active=True,
                    )
                    session.add(unit)
                    inserted_units += 1
                else:
                    if unit.unit_no != payload["unit_no"]:
                        unit.unit_no = payload["unit_no"]
                        renamed_units += 1
                    changed = unit.source_hash != payload["source_hash"]
                    for field in mutable_unit_fields:
                        setattr(unit, field, payload.get(field))
                    unit.source_hash = payload["source_hash"]
                    unit.updated_at = now
                    unit.last_seen_at = now
                    unit.is_active = True
                    if changed:
                        updated_units += 1
                    else:
                        unchanged_units += 1

            session.flush()
            deactivated_units = (
                session.query(Unit)
                .filter(Unit.unit_no.notin_(source_unit_numbers))
                .update(
                    {Unit.is_active: False, Unit.updated_at: now, Unit.last_seen_at: now},
                    synchronize_session=False,
                )
            )
            session.flush()

            existing_earnings = {row.receipt_key: row for row in session.query(Earning).all()}
            source_receipt_keys = {row["receipt_key"] for row in earnings}
            mutable_earning_fields = [
                column.name
                for column in Earning.__table__.columns
                if column.name not in {"earning_key", "receipt_key", "created_at", "first_seen_at", "updated_at", "last_seen_at", "is_active", "source_hash"}
            ]
            inserted_earnings = 0
            updated_earnings = 0
            unchanged_earnings = 0
            for row in earnings:
                source_hash = hash_row("earning-base-data", row)
                earning = existing_earnings.get(row["receipt_key"])
                if earning is None:
                    session.add(Earning(
                        **row,
                        source_hash=source_hash,
                        created_at=now,
                        updated_at=now,
                        first_seen_at=now,
                        last_seen_at=now,
                        is_active=True,
                    ))
                    inserted_earnings += 1
                else:
                    changed = earning.source_hash != source_hash
                    for field in mutable_earning_fields:
                        setattr(earning, field, row.get(field))
                    earning.source_hash = source_hash
                    earning.updated_at = now
                    earning.last_seen_at = now
                    earning.is_active = True
                    if changed:
                        updated_earnings += 1
                    else:
                        unchanged_earnings += 1
            session.flush()
            stale_earnings = (
                session.query(Earning)
                .filter(Earning.receipt_key.notin_(source_receipt_keys))
                .delete(synchronize_session=False)
            )
            session.flush()
            session.add_all(
                [
                    EarningLink(
                        receipt_key=row["receipt_key"],
                        unit_no=row["unit_no"],
                        station_code=row["station_code"],
                        match_status="Matched" if row["unit_no"] else "Missing unit",
                    )
                    for row in earnings
                ]
            )
            session.add(
                DataChangeLog(
                    resource="catering",
                    record_key=None,
                    action="replace_import",
                    source="xlsx",
                    details=(
                        f"{len(units)} units and {len(earnings)} earnings imported from "
                        f"{source_name}; {report['reconciliation']['linked_earning_rows']} earnings linked"
                    ),
                    created_at=now,
                )
            )
            session.flush()
            transaction_totals = {
                "active_units": session.query(func.count(Unit.unit_no)).filter(Unit.is_active.is_(True)).scalar() or 0,
                "earnings": session.query(func.count(Earning.earning_key)).scalar() or 0,
                "earning_links": session.query(func.count(EarningLink.id)).scalar() or 0,
                "earning_amount_total": session.query(func.coalesce(func.sum(Earning.amount), 0)).scalar() or 0,
            }
            if transaction_totals != {
                "active_units": len(units),
                "earnings": len(earnings),
                "earning_links": len(earnings),
                "earning_amount_total": report["source"]["earning_amount_total"],
            }:
                raise RuntimeError(f"In-transaction verification failed: {transaction_totals}")

            report["applied"] = {
                "inserted_units": inserted_units,
                "updated_units": updated_units,
                "unchanged_units": unchanged_units,
                "renamed_units": renamed_units,
                "deactivated_units": deactivated_units,
                "inserted_earnings": inserted_earnings,
                "updated_earnings": updated_earnings,
                "unchanged_earnings": unchanged_earnings,
                "removed_stale_earnings": stale_earnings,
                "rebuilt_earning_links": len(earnings),
            }
            session.add(CateringSyncRun(
                source_spreadsheet_id=spreadsheet_id,
                status="success",
                started_at=now,
                completed_at=datetime.now(timezone.utc),
                unit_rows=len(units),
                earning_source_rows=report["source"].get("earning_source_rows", len(earnings)),
                earning_rows=len(earnings),
                duplicate_rows=report["source"].get("duplicate_earning_rows", 0),
                linked_earnings=report["reconciliation"]["linked_earning_rows"],
                unlinked_earnings=report["reconciliation"]["unlinked_earning_rows"],
                report_json=json.dumps(report, default=str),
            ))

    finally:
        session.close()


def verify(report: dict[str, Any]) -> None:
    with SessionLocal() as session:
        database = {
            "units": session.query(func.count(Unit.unit_no)).filter(Unit.is_active.is_(True)).scalar() or 0,
            "earnings": session.query(func.count(Earning.earning_key)).scalar() or 0,
            "earning_links": session.query(func.count(EarningLink.id)).scalar() or 0,
            "linked_earnings": session.query(func.count(EarningLink.id)).filter(EarningLink.unit_no.is_not(None)).scalar()
            or 0,
            "earning_amount_total": session.query(func.coalesce(func.sum(Earning.amount), 0)).scalar() or 0,
            "gst_total": session.query(func.coalesce(func.sum(Earning.gst), 0)).scalar() or 0,
        }
    expected = report["source"]
    if database["units"] != expected["units"]:
        raise RuntimeError(f"Verification failed: expected {expected['units']} active units, got {database['units']}")
    if database["earnings"] != expected["earnings"] or database["earning_links"] != expected["earnings"]:
        raise RuntimeError("Verification failed: earning or earning-link row counts differ from the source")
    if database["earning_amount_total"] != expected["earning_amount_total"]:
        raise RuntimeError("Verification failed: earning amount total differs from the source")
    report["verified"] = database


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate and import updated catering unit and earning base data.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--apply", action="store_true", help="Write to PostgreSQL. Without this flag, runs validation only.")
    args = parser.parse_args()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    if engine.dialect.name != "postgresql":
        raise RuntimeError(f"PostgreSQL is required; connected dialect is {engine.dialect.name!r}")

    units, raw_earnings, notes = load_source(workbook_path)
    earnings, report = reconcile(units, raw_earnings)
    report["workbook"] = str(workbook_path)
    report["source"].update(notes)
    report["mode"] = "apply" if args.apply else "dry-run"
    if args.apply:
        apply_import(workbook_path.name, units, earnings, report)
        verify(report)
    print(json.dumps(report, indent=2, default=str))


if __name__ == "__main__":
    main()
