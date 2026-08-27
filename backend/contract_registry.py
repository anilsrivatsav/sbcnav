from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy import func, or_

from database import SessionLocal
from models import (
    ContractRegistryAsset,
    ContractRegistryContract,
    ContractRegistryContractor,
    ContractRegistryPaymentSchedule,
    ContractRegistryPayment,
    ContractRegistryStatusHistory,
    CommercialContract,
    Unit,
    Station,
)


def _text(value: Any) -> str | None:
    if value is None:
        return None
    value = str(value).replace("\xa0", " ").strip()
    return value or None


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text or text.startswith("#"):
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None


def _money(value: Any) -> float | None:
    try:
        if value is None or str(value).strip().startswith("#"):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalized(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def _status(value: Any) -> str:
    text = (_text(value) or "unknown").lower()
    if "cancel" in text:
        return "cancelled"
    if "terminat" in text:
        return "terminated"
    if "complete" in text or "closed" in text:
        return "completed"
    if "run" in text or "active" in text:
        return "running"
    if "award" in text:
        return "awarded"
    return "unknown"


def _family(policy: str | None, category: str | None, asset: str | None) -> str:
    text = " ".join(x or "" for x in (policy, category, asset)).lower()
    if "audio" in text:
        return "audio"
    if "train" in text or "demu" in text or "vande" in text or "mobile" in text or "rake" in text:
        return "mobile_asset"
    if "parking" in text:
        return "parking"
    if "advert" in text or "ooh" in text:
        return "advertisement"
    if any(x in text for x in ("station", "rdn", "nfris", "atm", "massage", "locker", "nursing")):
        return "station"
    return "other"


def _asset_type(asset: str | None, family: str) -> str:
    text = (asset or "").lower()
    if re.search(r"\d{4,6}[/-]\d{4,6}", text) or family == "mobile_asset":
        return "train"
    if family == "audio":
        return "audio"
    if family in {"advertisement", "parking", "station"}:
        return "station"
    return "other"


def parse_eauction_workbook(path_or_bytes: str | bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to import contracts") from exc
    source = path_or_bytes
    if isinstance(path_or_bytes, bytes):
        import io
        source = io.BytesIO(path_or_bytes)
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["E Auction"] if "E Auction" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_text(value) or "" for value in rows[0]]
    normalized_headers = [_normalized(value) for value in headers]
    # The live E Auction register no longer has the old Asset column.  Keep the
    # legacy layout readable, but map the current sheet explicitly so a column
    # insertion/removal cannot silently shift every field again.
    current_layout = len(normalized_headers) >= 17 and normalized_headers[3] == "policy" and normalized_headers[4] == "contract date"
    columns = ({
        "sl": 1, "contractor": 2, "number": 3, "asset": None,
        "policy": 4, "contract_date": 5, "category": 6,
        "annual": 7, "total": 8, "additional": 9, "status": 10,
        "quarterly": 11, "start": 12, "end": 13, "duration": 14,
        # P is a days-to-due formula. Actual installment dates start at Q.
        "schedule_start": 17, "schedule_end": 27,
    } if current_layout else {
        "sl": 1, "contractor": 2, "number": 3, "asset": 4,
        "policy": 5, "contract_date": 6, "category": 7,
        "annual": 8, "total": 9, "additional": 10, "status": 11,
        "quarterly": 12, "start": 13, "end": 14, "duration": 15,
        "schedule_start": 17, "schedule_end": 28,
    })
    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        value = lambda col: row[col - 1] if col - 1 < len(row) else None
        raw_number = value(columns["number"])
        raw_contractor = value(columns["contractor"])
        number = _text(raw_number)
        contractor = _text(raw_contractor)
        # Contract references in this register are textual (for example
        # SBC-OOH-VB-2022). Numeric-only footer totals must not become records.
        if not isinstance(raw_number, str) or not isinstance(raw_contractor, str) or not number or not contractor or number.lower() in {"contract number", "total"}:
            continue
        asset_col = columns["asset"]
        policy = _text(value(columns["policy"]))
        category = _text(value(columns["category"]))
        asset = _text(value(asset_col)) if asset_col else None
        status = _status(value(columns["status"]))
        annual_fee = _money(value(columns["annual"]))
        quarterly_fee = _money(value(columns["quarterly"]))
        if quarterly_fee is None and annual_fee is not None:
            quarterly_fee = annual_fee / 4
        schedules = []
        for col in range(columns["schedule_start"], min(len(row), columns["schedule_end"]) + 1):
            due = _date(value(col))
            if due:
                installment = col - columns["schedule_start"] + 1
                schedules.append({
                    "installment_number": installment,
                    "period_label": f"Payment {installment}",
                    "due_date": due,
                    "expected_amount": quarterly_fee,
                    "status": "past_due_date" if due < date.today() else "upcoming",
                })
        records.append({
            "source_sl_no": int(value(columns["sl"])) if isinstance(value(columns["sl"]), (int, float)) else None,
            "source_sheet": sheet.title, "source_row_number": row_number,
            "contract_number": number, "contract_name": category or number,
            "contractor": contractor, "asset": asset, "policy_code": policy, "category": category,
            "status": status, "contract_date": _date(value(columns["contract_date"])), "period_start": _date(value(columns["start"])),
            "period_end": _date(value(columns["end"])), "duration_value": _money(value(columns["duration"])),
            "annual_license_fee": annual_fee, "total_contract_value": _money(value(columns["total"])),
            "additional_license_fee": _money(value(columns["additional"])), "quarterly_license_fee": quarterly_fee,
            "schedules": schedules,
        })
    return records


def import_eauction_workbook(path_or_bytes: str | bytes, source_file: str | None = None) -> dict[str, int]:
    records = parse_eauction_workbook(path_or_bytes)
    if not records:
        raise ValueError("E Auction sheet contains no valid contract rows")
    numbers = [item["contract_number"] for item in records]
    if len(numbers) != len(set(numbers)):
        raise ValueError("E Auction sheet contains duplicate contract numbers")
    session = SessionLocal()
    created = updated = deleted = assets = schedules = statuses = 0
    try:
        station_codes = {row[0].upper() for row in session.query(Station.station_code).all()}
        for item in records:
            normalized = _normalized(item["contractor"])
            contractor = session.query(ContractRegistryContractor).filter_by(normalized_name=normalized).one_or_none()
            if not contractor:
                contractor = ContractRegistryContractor(legal_name=item["contractor"], normalized_name=normalized)
                session.add(contractor); session.flush(); created += 1
            contract = session.query(ContractRegistryContract).filter_by(source_system="e_auction", contract_number=item["contract_number"]).one_or_none()
            family = _family(item["policy_code"], item["category"], item["asset"])
            payload = dict(source_sl_no=item["source_sl_no"], source_system="e_auction", source_file=source_file or "workbook", source_sheet=item["source_sheet"], source_row_number=item["source_row_number"], contract_number=item["contract_number"], contract_name=item["contract_name"], contractor_id=contractor.contractor_id, contract_family=family, award_method="e_auction", policy_code=item["policy_code"], category=item["category"], status=item["status"], contract_date=item["contract_date"], loa_date=item["contract_date"], commencement_date=item["period_start"], period_start=item["period_start"], period_end=item["period_end"], duration_value=item["duration_value"], duration_unit="years", annual_license_fee=item["annual_license_fee"], quarterly_license_fee=item["quarterly_license_fee"], total_contract_value=item["total_contract_value"], additional_license_fee=item["additional_license_fee"])
            if not contract:
                contract = ContractRegistryContract(**payload); session.add(contract); session.flush(); created += 1
            else:
                for key, value in payload.items(): setattr(contract, key, value)
                updated += 1
            asset_value = item["asset"]
            if asset_value:
                session.query(ContractRegistryAsset).filter_by(contract_id=contract.contract_id).delete(synchronize_session=False)
                atype = _asset_type(asset_value, family)
                station_code = asset_value.strip().upper() if asset_value.strip().upper() in station_codes else None
                session.add(ContractRegistryAsset(contract_id=contract.contract_id, asset_type=atype, station_code=station_code, train_number=asset_value if atype == "train" else None, asset_name=asset_value, raw_asset_value=asset_value, match_status="matched" if station_code else "unmatched")); assets += 1
            session.query(ContractRegistryPaymentSchedule).filter_by(contract_id=contract.contract_id).delete(synchronize_session=False)
            for schedule in item["schedules"]:
                session.add(ContractRegistryPaymentSchedule(contract_id=contract.contract_id, **schedule)); schedules += 1
            if not session.query(ContractRegistryStatusHistory).filter_by(contract_id=contract.contract_id, status=item["status"]).first():
                session.add(ContractRegistryStatusHistory(contract_id=contract.contract_id, status=item["status"], effective_from=item["period_start"], source_reference=f"{item['source_sheet']}:{item['source_row_number']}")); statuses += 1
        stale = session.query(ContractRegistryContract).filter(
            ContractRegistryContract.source_system == "e_auction",
            ContractRegistryContract.contract_number.notin_(numbers),
        ).all()
        for contract in stale:
            session.delete(contract)
            deleted += 1
        session.commit()
        return {"records": len(records), "created": created, "updated": updated, "deleted": deleted, "assets": assets, "schedules": schedules, "statuses": statuses}
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def backfill_legacy_contracts() -> dict[str, int]:
    """Expose the app's existing commercial/catering contracts in the unified registry."""
    session = SessionLocal()
    created = 0
    try:
        rows = []
        for row in session.query(CommercialContract).all():
            status = "cancelled" if row.termination_status and "cancel" in row.termination_status.lower() else "running"
            rows.append({"source_system": "legacy_commercial", "source_number": f"commercial-{row.contract_key}", "name": row.contract_name, "contractor": row.licensee_name, "family": row.asset_scope or "other", "policy": row.policy, "category": row.sub_category, "status": status, "start": _date(row.contract_period_from), "end": _date(row.contract_upto), "annual": row.annual_license_fee, "quarterly": row.quarterly_license_fee, "total": None, "asset": row.raw_station_value})
        for row in session.query(Unit).filter(Unit.licensee_name.isnot(None)).all():
            rows.append({"source_system": "legacy_catering", "source_number": f"unit-{row.unit_no}", "name": row.type_of_unit or row.unit_no, "contractor": row.licensee_name, "family": "catering", "policy": row.allotment_type, "category": row.type_of_unit, "status": "running" if row.contract_to else "unknown", "start": _date(row.contract_from), "end": _date(row.contract_to), "annual": _money(row.license_fee), "quarterly": None, "total": None, "asset": row.station_code})
        for item in rows:
            normalized = _normalized(item["contractor"] or "unknown contractor")
            contractor = session.query(ContractRegistryContractor).filter_by(normalized_name=normalized).one_or_none()
            if not contractor:
                contractor = ContractRegistryContractor(legal_name=item["contractor"] or "Unknown contractor", normalized_name=normalized); session.add(contractor); session.flush()
            contract = session.query(ContractRegistryContract).filter_by(source_system=item["source_system"], contract_number=item["source_number"]).one_or_none()
            payload = dict(source_system=item["source_system"], contract_number=item["source_number"], contract_name=item["name"], contractor_id=contractor.contractor_id, contract_family=item["family"], award_method="legacy_import", policy_code=item["policy"], category=item["category"], status=item["status"], period_start=item["start"], period_end=item["end"], annual_license_fee=item["annual"], quarterly_license_fee=item["quarterly"], total_contract_value=item["total"], source_file="existing rail inspect tables")
            if not contract:
                contract = ContractRegistryContract(**payload); session.add(contract); session.flush(); created += 1
            else:
                for key, value in payload.items(): setattr(contract, key, value)
            session.query(ContractRegistryAsset).filter_by(contract_id=contract.contract_id).delete(synchronize_session=False)
            if item["asset"]:
                raw = str(item["asset"]); station_code = raw.strip().upper() if raw.strip().upper() in {row[0].upper() for row in session.query(Station.station_code).all()} else None
                session.add(ContractRegistryAsset(contract_id=contract.contract_id, asset_type="station" if station_code else "other", station_code=station_code, asset_name=raw, raw_asset_value=raw, match_status="matched" if station_code else "unmatched"))
        session.commit()
        return {"legacy_records": len(rows), "legacy_created": created}
    except Exception:
        session.rollback(); raise
    finally:
        session.close()


def _contract_dict(row: ContractRegistryContract, contractor: ContractRegistryContractor | None, assets: list[ContractRegistryAsset], schedules: list[ContractRegistryPaymentSchedule], payments: list[ContractRegistryPayment] | None = None) -> dict[str, Any]:
    payment_rows = payments or []
    payment_items = [{"payment_id": item.payment_id, "schedule_id": item.schedule_id, "payment_date": item.payment_date, "amount_due": item.amount_due, "amount_paid": item.amount_paid, "interest_amount": item.interest_amount, "delay_days": item.delay_days, "payment_reference": item.payment_reference, "payment_status": item.payment_status, "source_reference": item.source_reference, "remarks": item.remarks} for item in payment_rows]
    return {"contract_id": row.contract_id, "contract_number": row.contract_number, "contract_name": row.contract_name, "status": row.status, "contract_family": row.contract_family, "award_method": row.award_method, "policy_code": row.policy_code, "category": row.category, "period": {"loa_date": row.loa_date, "start": row.period_start, "end": row.period_end, "duration_value": row.duration_value, "duration_unit": row.duration_unit}, "financials": {"annual_license_fee": row.annual_license_fee, "quarterly_license_fee": row.quarterly_license_fee, "total_contract_value": row.total_contract_value, "additional_license_fee": row.additional_license_fee}, "contractor": {"contractor_id": contractor.contractor_id, "legal_name": contractor.legal_name} if contractor else None, "assets": [{"asset_type": item.asset_type, "station_code": item.station_code, "train_number": item.train_number, "asset_name": item.asset_name, "raw_asset_value": item.raw_asset_value, "match_status": item.match_status} for item in assets], "payment_schedule": [{"schedule_id": item.schedule_id, "installment_number": item.installment_number, "period_label": item.period_label, "due_date": item.due_date, "expected_amount": item.expected_amount, "status": item.status} for item in schedules], "payments": payment_items, "payment_summary": {"scheduled": len(schedules), "scheduled_amount": sum(float(item.expected_amount or 0) for item in schedules), "recorded": len(payment_items), "amount_due": sum(float(item.amount_due or 0) for item in payment_rows), "amount_paid": sum(float(item.amount_paid or 0) for item in payment_rows), "pending": sum(1 for item in payment_rows if str(item.payment_status or '').lower() in {'pending', 'overdue'})}, "source": {"system": row.source_system, "file": row.source_file, "sheet": row.source_sheet, "row": row.source_row_number}}


def list_registry_contracts(status: str | None = None, search: str | None = None, asset_type: str | None = None) -> list[dict[str, Any]]:
    session = SessionLocal()
    try:
        query = session.query(ContractRegistryContract, ContractRegistryContractor).join(ContractRegistryContractor, ContractRegistryContractor.contractor_id == ContractRegistryContract.contractor_id, isouter=True)
        if status and status != "all": query = query.filter(ContractRegistryContract.status == status)
        if search:
            like = f"%{search}%"
            query = query.filter(or_(ContractRegistryContract.contract_number.ilike(like), ContractRegistryContract.contract_name.ilike(like), ContractRegistryContract.policy_code.ilike(like), ContractRegistryContractor.legal_name.ilike(like)))
        contract_rows = query.order_by(ContractRegistryContract.status, ContractRegistryContract.period_end, ContractRegistryContract.contract_name).all()
        contract_ids = [contract.contract_id for contract, _ in contract_rows]
        assets_by_contract: dict[int, list[ContractRegistryAsset]] = {}
        schedules_by_contract: dict[int, list[ContractRegistryPaymentSchedule]] = {}
        payments_by_contract: dict[int, list[ContractRegistryPayment]] = {}
        if contract_ids:
            for item in session.query(ContractRegistryAsset).filter(ContractRegistryAsset.contract_id.in_(contract_ids)).all():
                assets_by_contract.setdefault(item.contract_id, []).append(item)
            for item in session.query(ContractRegistryPaymentSchedule).filter(ContractRegistryPaymentSchedule.contract_id.in_(contract_ids)).order_by(ContractRegistryPaymentSchedule.installment_number).all():
                schedules_by_contract.setdefault(item.contract_id, []).append(item)
            for item in session.query(ContractRegistryPayment).filter(ContractRegistryPayment.contract_id.in_(contract_ids)).order_by(ContractRegistryPayment.payment_date, ContractRegistryPayment.payment_id).all():
                payments_by_contract.setdefault(item.contract_id, []).append(item)
        rows = []
        for contract, contractor in contract_rows:
            assets = assets_by_contract.get(contract.contract_id, [])
            if asset_type and asset_type != "all" and not any(item.asset_type == asset_type for item in assets):
                continue
            rows.append(_contract_dict(contract, contractor, assets, schedules_by_contract.get(contract.contract_id, []), payments_by_contract.get(contract.contract_id, [])))
        return rows
    finally: session.close()


def get_registry_contract(contract_id: int) -> dict[str, Any] | None:
    session = SessionLocal()
    try:
        row = session.get(ContractRegistryContract, contract_id)
        if not row: return None
        contractor = session.get(ContractRegistryContractor, row.contractor_id) if row.contractor_id else None
        assets = session.query(ContractRegistryAsset).filter_by(contract_id=contract_id).all()
        schedules = session.query(ContractRegistryPaymentSchedule).filter_by(contract_id=contract_id).order_by(ContractRegistryPaymentSchedule.installment_number).all()
        payments = session.query(ContractRegistryPayment).filter_by(contract_id=contract_id).order_by(ContractRegistryPayment.payment_date, ContractRegistryPayment.payment_id).all()
        return _contract_dict(row, contractor, assets, schedules, payments)
    finally: session.close()


def registry_summary() -> dict[str, Any]:
    session = SessionLocal()
    try:
        counts = {status: session.query(ContractRegistryContract).filter(ContractRegistryContract.status == status).count() for status in ("awarded", "running", "completed", "cancelled", "terminated", "unknown")}
        return {"counts": counts, "total": sum(counts.values()), "total_contract_value": float(session.query(func.coalesce(func.sum(ContractRegistryContract.total_contract_value), 0)).scalar() or 0)}
    finally: session.close()
