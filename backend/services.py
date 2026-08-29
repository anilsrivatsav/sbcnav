from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import func, select

from database import SessionLocal
from models import (
    AmenityNorm,
    CommercialContract,
    CommercialContractPayment,
    CommercialContractStationLink,
    CateringSyncRun,
    DataChangeLog,
    Earning,
    EarningLink,
    PassengerAmenityWork,
    PlatformExtensionSummary,
    PlatformDetail,
    Inspection,
    InspectionFinding,
    MobileDeviceState,
    Station,
    StationMonthlyMetric,
    StationInfra,
    StationPlatformExtensionStatus,
    TrolleyPath,
    Unit,
    WheelChairAvailability,
    Work,
    WorkProgressUpdate,
    WorkExpenditureUpdate,
    WorkLink,
)
from contract_registry import list_registry_contracts

def clean(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return "" if not text or text.upper() == "#N/A" else text


def to_int(value: Any) -> int | None:
    text = clean(value).replace("₹", "").replace("?", "").replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def to_money(value: Any) -> int:
    parsed = to_int(value)
    return parsed or 0


def parse_date_value(value: Any) -> date | None:
    text = clean(value)
    if not text:
        return None
    text = re.sub(r"\s+00:00:00$", "", text)
    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d.%m.%Y",
        "%Y/%m/%d",
        "%d-%b-%Y",
        "%d %b %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", text)
    if match:
        day, month, year = match.groups()
        year_int = int(year)
        if year_int < 100:
            year_int += 2000
        try:
            return date(year_int, int(month), int(day))
        except ValueError:
            return None
    return None


def date_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    parsed = parse_date_value(value)
    return parsed.isoformat() if parsed else clean(value) or None


def month_end(value: date) -> date:
    if value.month == 12:
        return date(value.year, 12, 31)
    return date(value.year, value.month + 1, 1) - timedelta(days=1)


def month_delta(start: date, end: date) -> int:
    if start > end:
        return 0
    return (end.year - start.year) * 12 + end.month - start.month + 1


def is_active_status(value: Any) -> bool:
    text = normalize(clean(value))
    if not text:
        return True
    return not any(token in text for token in ["available", "closed", "inactive", "terminated", "expired", "vacant", "surrender"])


def is_available_unit(unit: Unit | dict[str, Any]) -> bool:
    def value(key: str) -> Any:
        return unit.get(key) if isinstance(unit, dict) else getattr(unit, key, None)

    if normalize(value("unit_status")) == "available":
        return True
    return not clean(value("licensee_name")) and not clean(value("contract_from")) and not clean(value("contract_to"))


def contract_risk(
    valid_to: Any,
    *,
    pending_amount: int = 0,
    station_match_status: Any = None,
    missing_fee: bool = False,
    today: date | None = None,
) -> dict[str, Any]:
    """Return one consistent risk/notification shape for every contract source."""
    today = today or date.today()
    expiry = parse_date_value(valid_to)
    days = (expiry - today).days if expiry else None
    if days is None:
        expiry_points = 25
        bucket = "no_date"
        renewal_state = "Date unavailable"
    elif days < 0:
        expiry_points = 100
        bucket = "expired"
        renewal_state = "Expired"
    elif days == 0:
        expiry_points = 95
        bucket = "due_today"
        renewal_state = "Due today"
    elif days <= 5:
        expiry_points = 90
        bucket = "due_5_days"
        renewal_state = f"Due within {days} days"
    elif days <= 10:
        expiry_points = 80
        bucket = "due_10_days"
        renewal_state = f"Due within {days} days"
    elif days <= 30:
        expiry_points = 60
        bucket = "due_30_days"
        renewal_state = f"Due within {days} days"
    elif days <= 50:
        expiry_points = 45
        bucket = "due_50_days"
        renewal_state = f"Due within {days} days"
    elif days <= 90:
        expiry_points = 35
        bucket = "due_90_days"
        renewal_state = f"Due within {days} days"
    else:
        expiry_points = 10
        bucket = "over_90_days"
        renewal_state = "Active"

    match_text = normalize(station_match_status)
    risk_score = min(
        100,
        expiry_points
        + (20 if pending_amount > 0 else 0)
        + (10 if missing_fee else 0)
        + (10 if match_text in {"", "unmatched", "asset_scope", "missing link"} else 0),
    )
    risk_level = "critical" if risk_score >= 80 else "high" if risk_score >= 60 else "medium" if risk_score >= 35 else "low"
    return {
        "valid_to": expiry.isoformat() if expiry else clean(valid_to) or None,
        "days_to_expiry": days,
        "expiry_alert_bucket": bucket,
        "renewal_state": renewal_state,
        "pending_amount": pending_amount,
        "risk_score": risk_score,
        "risk_level": risk_level,
        "notification_required": days is not None and days <= 90 or pending_amount > 0 or missing_fee,
    }


def is_license_fee_row(row: Earning) -> bool:
    text = normalize(" ".join([clean(row.payment_head), clean(row.payment_sub_head), clean(row.receipt_type)]))
    return "license" in text or "licence" in text or "lf" == text


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", clean(text).lower())


def parse_csv(text: str) -> list[list[str]]:
    return list(csv.reader(io.StringIO(text)))


def hash_row(prefix: str, payload: dict[str, Any]) -> str:
    material = prefix + "|" + "|".join(f"{k}={clean(v)}" for k, v in sorted(payload.items()))
    return hashlib.sha256(material.encode("utf-8")).hexdigest()


def audit_fields(now: datetime) -> dict[str, Any]:
    return {
        "created_at": now,
        "updated_at": now,
        "first_seen_at": now,
        "last_seen_at": now,
        "is_active": True,
    }


def parse_stations(text: str) -> list[dict[str, Any]]:
    rows = parse_csv(text)
    headers = [normalize(h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        item: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else ""
            mapping = {
                "station code": "station_code",
                "station name": "station_name",
                "division": "division",
                "zone": "zone",
                "section": "section",
                "cmi": "cmi",
                "den": "den",
                "sr.den": "sr_den",
                "sr den": "sr_den",
                "sr den name": "sr_den",
                "categorisation": "categorisation",
                "earnings range": "earnings_range",
                "passenger range": "passenger_range",
                "passenger footfall": "passenger_footfall",
                "platforms": "platforms",
                "number of platforms": "number_of_platforms",
                "platform type": "platform_type",
                "parking": "parking",
                "pay-and-use": "pay_and_use",
                "no of trains dealt": "trains_dealt",
                "tkts per day": "tickets_per_day",
                "pass per day": "passengers_per_day",
                "earnings per day": "earnings_per_day",
                "footfalls per day": "footfalls_per_day",
            }
            if header in mapping:
                item[mapping[header]] = to_int(value) if header in {"passenger footfall", "number of platforms", "no of trains dealt", "tkts per day", "pass per day", "earnings per day", "footfalls per day"} else clean(value)
        if item.get("station_code"):
            out.append(item)
    return out


def parse_units(text: str) -> list[dict[str, Any]]:
    rows = parse_csv(text)
    headers = [normalize(h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        item: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else ""
            mapping = {
                "sl no.": "sl_no",
                "unit no.": "unit_no",
                "type of unit": "type_of_unit",
                "station": "station_code",
                "station category": "station_category",
                "old category": "old_category",
                "pf no": "pf_no",
                "pegged location": "pegged_location",
                "reservation category": "reservation_category",
                "type of allotment": "allotment_type",
                "name of licensee": "licensee_name",
                "license fee": "license_fee",
                "contract from": "contract_from",
                "contract to": "contract_to",
                "unit status": "unit_status",
            }
            if header in mapping:
                item[mapping[header]] = to_int(value) if header == "sl no." else clean(value).replace("₹", "")
        if item.get("unit_no"):
            out.append(item)
    return out


def parse_earnings(text: str) -> list[dict[str, Any]]:
    rows = parse_csv(text)
    headers = [normalize(h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        item: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else ""
            mapping = {
                "sl. no.": "sl_no",
                "date of receipt": "date_of_receipt",
                "unit no.": "unit_no",
                "station": "station_code",
                "pf no.": "pf_no",
                "name of licensee": "licensee_name",
                "payment head": "payment_head",
                "payment sub-head": "payment_sub_head",
                "period from": "period_from",
                "period to": "period_to",
                "amount": "amount",
                "gst": "gst",
                "reciept type": "receipt_type",
                "mr no/uts no/ challan no": "mr_no",
                "mr date": "mr_date",
                "u/a case": "ua_case",
            }
            if header in mapping:
                item[mapping[header]] = to_int(value) if header in {"sl. no.", "amount", "gst"} else clean(value)
        if item.get("unit_no"):
            item["receipt_key"] = hash_row("earning", item)
            out.append(item)
    return out


def parse_works(text: str) -> list[dict[str, Any]]:
    rows = parse_csv(text)
    header_idx = next(i for i, row in enumerate(rows) if any(normalize(cell) == "projectid" for cell in row))
    headers = [normalize(h) for h in rows[header_idx]]
    parsed: list[dict[str, Any]] = []
    last_sn = 0
    for row in rows[header_idx + 1 :]:
        # The worksheet contains a second supplementary table below the
        # 152-row sanctioned-works register. It is not part of the total.
        if any(normalize(cell) == "projectid" for cell in row):
            break
        item: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else ""
            mapping = {
                "sn": "source_sn",
                "projectid": "project_id",
                "year of sanction": "year_of_sanction",
                "year ub works": "year_ub_works",
                "status": "status",
                "date of sanction": "date_of_sanction",
                "short name of work": "short_name_of_work",
                "block section station": "block_section_station",
                "allocation": "allocation",
                "cost": "cost",
                "expenditure upto date": "expenditure_upto_date",
                "physical progress in %": "physical_progress",
                "financial progress in %": "financial_progress",
                "financial progress": "financial_progress",
                "if ub?": "if_ub",
                "parent work": "parent_work",
                "section": "section",
                "remarks": "remarks",
            }
            if header.startswith("engg remarks"):
                item["engg_remarks"] = clean(value)
            elif header.startswith("anticipated expenditure"):
                item["anticipated_expenditure"] = to_int(value)
            elif header in mapping:
                target = mapping[header]
                item[target] = to_int(value) if target in {"source_sn", "cost", "expenditure_upto_date"} else clean(value)
        if item.get("source_sn"):
            if last_sn and item["source_sn"] <= last_sn:
                break
            last_sn = item["source_sn"]
        if item.get("project_id") and normalize(item["project_id"]) != "projectid":
            item["project_id"] = re.sub(r"\s+", "", clean(item["project_id"]))
            item["source_project_id"] = item["project_id"]
            parsed.append(item)

    counts: dict[str, int] = {}
    for item in parsed:
        counts[item["source_project_id"]] = counts.get(item["source_project_id"], 0) + 1
    for index, item in enumerate(parsed, start=1):
        if counts.get(item["source_project_id"], 0) > 1:
            suffix = item.get("source_sn") or index
            item["project_id"] = f"{item['source_project_id']}__SN{suffix}"
    return parsed


def parse_works_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Parse the complete sanctioned-works worksheet from an XLSX export.

    The Google Sheets CSV export can reflect the current sheet view/filter and
    is therefore not a reliable import source for this register. XLSX keeps
    the complete first table, including numeric cells and the 152-row range.
    The existing CSV parser remains the canonical field mapping/parser.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook["All Sanctioned Works"]
    rows = list(worksheet.iter_rows(values_only=True))
    header_idx = next(
        i for i, row in enumerate(rows)
        if any(normalize(cell) == "projectid" for cell in row)
    )

    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows[header_idx:]:
        if row and any(normalize(cell) == "projectid" for cell in row) and row is not rows[header_idx]:
            break
        converted = []
        for value in row:
            if isinstance(value, (datetime, date)):
                converted.append(value.strftime("%Y-%m-%d"))
            else:
                converted.append("" if value is None else str(value))
        writer.writerow(converted)
    return parse_works(output.getvalue())


def _header_map(headers: list[str]) -> dict[str, int]:
    return {normalize(header): index for index, header in enumerate(headers) if normalize(header)}


def _cell(row: list[str], indexes: dict[str, int], *names: str) -> str:
    for name in names:
        idx = indexes.get(normalize(name))
        if idx is not None and idx < len(row):
            value = clean(row[idx])
            if value:
                return value
    return ""


def parse_amenity_norms(text: str) -> list[dict[str, Any]]:
    rows = parse_csv(text)
    if not rows:
        return []
    indexes = _header_map(rows[0])
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {
            "category": _cell(row, indexes, "categorization"),
            "amenity": _cell(row, indexes, "amenity"),
            "norm": _cell(row, indexes, "norms"),
            "norm_quantity": _cell(row, indexes, "norms quantity"),
        }
        if item["category"] and item["norm"]:
            out.append(item)
    return out


def parse_station_infra(text: str) -> list[dict[str, Any]]:
    rows = parse_csv(text)
    if not rows:
        return []
    indexes = _header_map(rows[0])
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {
            "station_name": _cell(row, indexes, "stn"),
            "station_code": _cell(row, indexes, "code").upper(),
            "category": _cell(row, indexes, "catg"),
            "platform_list": clean(row[4] if len(row) > 4 else ""),
            "platform_count": to_int(_cell(row, indexes, "available")),
            "platform_level": _cell(row, indexes, "available"),
            "fob_details": _cell(row, indexes, "fob"),
            "shelter_details": _cell(row, indexes, "shelter"),
            "remarks": " | ".join(part for part in [_cell(row, indexes, "column 1"), _cell(row, indexes, "column 2"), _cell(row, indexes, "column 3")] if part),
        }
        count_candidate = to_int(row[5] if len(row) > 5 else "")
        level_candidate = clean(row[6] if len(row) > 6 else "")
        if count_candidate is not None:
            item["platform_count"] = count_candidate
        if level_candidate:
            item["platform_level"] = level_candidate
        if item["station_code"]:
            out.append(item)
    return out


def parse_platform_details(text: str) -> list[dict[str, Any]]:
    rows = parse_csv(text)
    if not rows:
        return []
    indexes = _header_map(rows[0])
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {
            "station_code": _cell(row, indexes, "station").upper(),
            "platform": _cell(row, indexes, "platforms"),
            "length_m": to_int(_cell(row, indexes, "length")),
            "lifts": _cell(row, indexes, "lifts"),
            "escalators": _cell(row, indexes, "escalators"),
            "ramp": _cell(row, indexes, "ramp"),
        }
        if item["station_code"] and item["platform"]:
            out.append(item)
    return out


def parse_combined_accessibility(text: str) -> list[dict[str, Any]]:
    """Parse the station-wise combined lift, ramp and escalator sheet."""
    rows = parse_csv(text)
    header_idx = next(
        (idx for idx, row in enumerate(rows) if normalize(_cell(row, _header_map(row), "station code")) == "station code"),
        1 if len(rows) > 1 else 0,
    )
    if header_idx >= len(rows):
        return []
    indexes = _header_map(rows[header_idx])
    out: list[dict[str, Any]] = []
    for source_row, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        station_code = _cell(row, indexes, "station code").upper()
        if not station_code or station_code in {"TOTAL", "NA", "NIL"}:
            continue
        lift = _cell(row, indexes, "lift")
        ramp = _cell(row, indexes, "ramp")
        escalator = _cell(row, indexes, "escalator")
        status_parts = [
            part for part in (
                "Lift: " + lift if lift else "",
                "Ramp: " + ramp if ramp else "",
                "Escalator: " + escalator if escalator else "",
            ) if part
        ]
        out.append({
            "station_code": station_code,
            "category": _cell(row, indexes, "category"),
            "source_category": _cell(row, indexes, "category"),
            "pf_extension_wip": False,
            "pf_extension_proposed": False,
            "raising_extension_proposed": False,
            "platform_extension_work_proposed": False,
            "ramp_feasible": False,
            "fob_without": False,
            "fob_ramp_available": "ramp" in ramp.lower() and "available" in ramp.lower(),
            "fob_wip": False,
            "lift_available": "available" in lift.lower() and "proposed" not in lift.lower(),
            "lift_proposed": "proposed" in lift.lower() or "sanctioned" in lift.lower() or "wip" in lift.lower(),
            "ramp_proposed": "proposed" in ramp.lower(),
            "not_feasible_lift_ramp": "not feasible" in (lift + " " + ramp).lower(),
            "footfall_day": to_int(_cell(row, indexes, "footfall / day", "footfall")),
            "lift_details": lift,
            "ramp_details": ramp,
            "escalator_details": escalator,
            "accessibility_source": "Combined lift, escalator and ramp sheet",
            "source_rows": str(source_row),
            "status_text": " | ".join(status_parts),
            "remarks": "Station-wise accessibility details from the latest combined sheet.",
        })
    return out


def parse_wheel_chairs(text: str) -> list[dict[str, Any]]:
    rows = parse_csv(text)
    header_idx = next((idx for idx, row in enumerate(rows) if any(normalize(cell) == "station code" for cell in row)), 0)
    indexes = _header_map(rows[header_idx])
    out: list[dict[str, Any]] = []
    for row in rows[header_idx + 1:]:
        item = {
            "station_code": _cell(row, indexes, "station code").upper(),
            "station_name": _cell(row, indexes, "station name"),
            "section": _cell(row, indexes, "section"),
            "category": _cell(row, indexes, "cat."),
            "available_good_condition": to_int(_cell(row, indexes, "no of wheel chairs available in good condition")),
        }
        if item["station_code"]:
            out.append(item)
    return out


def parse_trolley_paths(text: str) -> list[dict[str, Any]]:
    rows = parse_csv(text)
    if not rows:
        return []
    indexes = _header_map(rows[0])
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {
            "station_code": _cell(row, indexes, "station code").upper(),
            "station_name": _cell(row, indexes, "station name"),
            "division": _cell(row, indexes, "division"),
            "zone": _cell(row, indexes, "zone"),
            "section": _cell(row, indexes, "section"),
            "categorisation": _cell(row, indexes, "categorisation"),
            "passenger_footfall": to_int(_cell(row, indexes, "passenger footfall")),
            "platforms": _cell(row, indexes, "platforms"),
            "number_of_platforms": _cell(row, indexes, "number of platforms"),
            "platform_type": _cell(row, indexes, "platform type"),
            "trolley_path": _cell(row, indexes, "trolleypath"),
            "trolley_path_sanction": _cell(row, indexes, "trolleypath sanction"),
        }
        if item["station_code"]:
            out.append(item)
    return out


def parse_fob_works(text: str) -> list[dict[str, Any]]:
    rows = parse_csv(text)
    if not rows:
        return []
    indexes = _header_map(rows[0])
    out: list[dict[str, Any]] = []
    current_group = "FOB"
    for row in rows[1:]:
        if any(normalize(cell) == "abss" for cell in row):
            current_group = "FOB ABSS"
            continue
        station_code = _cell(row, indexes, "station").upper() or (clean(row[1]).upper() if len(row) > 1 else "")
        if not station_code or station_code in {"STATION", "ABSS"}:
            continue
        item = {
            "work_type": current_group,
            "station_code": station_code,
            "work_name": f"{current_group} work at {station_code}",
            "tender_status": _cell(row, indexes, "tender status"),
            "loa_date": _cell(row, indexes, "loa date"),
            "progress": _cell(row, indexes, "progress of work"),
            "physical_progress": _cell(row, indexes, "physical progress"),
        }
        out.append(item)
    return out


def parse_pf_extension_works(text: str, work_type: str = "PF Extension") -> list[dict[str, Any]]:
    rows = parse_csv(text)
    if not rows:
        return []
    headers = rows[0]
    out: list[dict[str, Any]] = []
    for row in rows[2:]:
        station_code = clean(row[3] if work_type == "HAS" and len(row) > 3 else row[1] if len(row) > 1 else "").upper()
        if not station_code:
            continue
        item = {
            "work_type": work_type,
            "project_id": clean(row[1]) if work_type == "HAS" and len(row) > 1 else None,
            "cost": clean(row[2]) if work_type == "HAS" and len(row) > 2 else None,
            "station_code": station_code,
            "station_category": clean(row[4] if work_type == "HAS" and len(row) > 4 else row[2] if len(row) > 2 else ""),
            "platform_level": clean(row[5] if work_type == "HAS" and len(row) > 5 else row[3] if len(row) > 3 else ""),
            "existing_platform_length": " | ".join(clean(row[idx]) for idx in ([6, 7, 8] if work_type == "HAS" else [4, 5, 6, 7]) if idx < len(row) and clean(row[idx])),
            "sanction_date": clean(row[9] if work_type == "HAS" and len(row) > 9 else row[8] if len(row) > 8 else ""),
            "executive_agency": clean(row[10] if work_type == "HAS" and len(row) > 10 else row[9] if len(row) > 9 else ""),
            "progress": clean(row[11] if work_type == "HAS" and len(row) > 11 else row[10] if len(row) > 10 else ""),
            "tdc": clean(row[12] if work_type == "HAS" and len(row) > 12 else row[11] if len(row) > 11 else ""),
            "work_name": f"{work_type} at {station_code}",
        }
        out.append(item)
    return out


STATION_CODE_STOPWORDS = {
    "ABSS", "AND", "AS", "AT", "BASIC", "CAT", "CATG", "DIV", "DIVISION", "EXISTING", "FEASIBLE",
    "FOB", "FOR", "FULL", "HAS", "LENGTH", "LIFT", "NIL", "NO", "NONE", "NOT", "OF", "PF",
    "PLATFORM", "PROPOSED", "RAMP", "RAISING", "REQUIRED", "STATION", "STATIONS", "TO", "TOTAL",
    "WIP", "WITH", "WITHOUT", "WORK", "WORKS",
}


def extract_station_codes(value: Any) -> list[str]:
    text = clean(value).upper()
    if not text:
        return []
    codes = []
    for token in re.findall(r"\b[A-Z]{1,5}[A-Z0-9]?\b", text):
        if token in STATION_CODE_STOPWORDS:
            continue
        if re.fullmatch(r"NSG|HG|SG|YES|WIP|NIL|NA", token):
            continue
        if len(token) < 2:
            continue
        codes.append(token)
    return list(dict.fromkeys(codes))


def _xlsx_cell(row, index: int) -> str:
    return clean(row[index].value if index < len(row) else "")


def _status_for(statuses: dict[str, dict[str, Any]], code: str) -> dict[str, Any]:
    return statuses.setdefault(code, {
        "station_code": code,
        "category": None,
        "source_category": None,
        "station_detail_category_code": None,
        "pf_extension_wip": False,
        "pf_extension_proposed": False,
        "raising_extension_proposed": False,
        "platform_extension_work_proposed": False,
        "ramp_feasible": False,
        "fob_without": False,
        "fob_ramp_available": False,
        "fob_wip": False,
        "lift_available": False,
        "lift_proposed": False,
        "ramp_proposed": False,
        "not_feasible_lift_ramp": False,
        "source_rows": "",
        "status_text": "",
        "remarks": "",
    })


def _append_status_text(row: dict[str, Any], field: str, text: str) -> None:
    value = clean(text)
    if not value:
        return
    existing = clean(row.get(field))
    parts = [part for part in existing.split(" | ") if part] if existing else []
    if value not in parts:
        parts.append(value)
    row[field] = " | ".join(parts)


def _mark_station_codes(statuses: dict[str, dict[str, Any]], source: Any, *, flag: str, category: str, source_row: int, label: str, remarks: str = "") -> None:
    for code in extract_station_codes(source):
        row = _status_for(statuses, code)
        row[flag] = True
        row["category"] = row["category"] or category
        row["source_category"] = row["source_category"] or category
        _append_status_text(row, "source_rows", str(source_row))
        _append_status_text(row, "status_text", f"{label}: {clean(source)}")
        _append_status_text(row, "remarks", remarks)


def parse_platform_extension_workbook(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to import platform extension workbook") from exc

    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=True)
    if "PF Extn and Raising" not in wb.sheetnames:
        raise ValueError("PF Extn and Raising sheet was not found in the workbook")
    ws = wb["PF Extn and Raising"]
    rows = list(ws.iter_rows())
    summaries: list[dict[str, Any]] = []
    statuses: dict[str, dict[str, Any]] = {}

    for row_number in range(5, 14):
        row = rows[row_number - 1]
        category = _xlsx_cell(row, 1)
        if not category or normalize(category) == "total":
            continue
        item = {
            "summary_type": "platform_extension",
            "category": category,
            "station_count": to_int(_xlsx_cell(row, 2)),
            "full_length_platforms": to_int(_xlsx_cell(row, 3)),
            "work_under_progress": _xlsx_cell(row, 4),
            "pf_extension_proposed": _xlsx_cell(row, 5),
            "raising_extension_proposed": _xlsx_cell(row, 6),
            "remarks": _xlsx_cell(row, 7),
            "platform_extension_work_proposed": _xlsx_cell(row, 8),
            "existing_length": _xlsx_cell(row, 9),
            "required_length": _xlsx_cell(row, 10),
            "source_row": row_number,
        }
        summaries.append(item)
        _mark_station_codes(statuses, item["work_under_progress"], flag="pf_extension_wip", category=category, source_row=row_number, label="PF extension WIP", remarks=item["remarks"])
        _mark_station_codes(statuses, item["pf_extension_proposed"], flag="pf_extension_proposed", category=category, source_row=row_number, label="PF extension proposed", remarks=item["remarks"])
        _mark_station_codes(statuses, item["raising_extension_proposed"], flag="raising_extension_proposed", category=category, source_row=row_number, label="Raising and extension proposed", remarks=item["remarks"])
        _mark_station_codes(statuses, item["platform_extension_work_proposed"], flag="platform_extension_work_proposed", category=category, source_row=row_number, label="Platform extension work proposed", remarks=item["remarks"])

    for row_number in range(19, 28):
        row = rows[row_number - 1]
        category = _xlsx_cell(row, 1)
        if not category or normalize(category) == "total":
            continue
        item = {
            "summary_type": "accessibility",
            "category": category,
            "station_count": to_int(_xlsx_cell(row, 2)),
            "fob_ramps_stairs_available": _xlsx_cell(row, 3),
            "stations_without_fob": _xlsx_cell(row, 4),
            "stations_with_fob_ramp": _xlsx_cell(row, 5),
            "stations_fob_wip": _xlsx_cell(row, 6),
            "stations_with_lift": _xlsx_cell(row, 7),
            "stations_lift_proposed": _xlsx_cell(row, 8),
            "stations_ramp_proposed": _xlsx_cell(row, 9),
            "stations_not_feasible_lift_ramp": _xlsx_cell(row, 10),
            "remarks": _xlsx_cell(row, 11),
            "source_row": row_number,
        }
        summaries.append(item)
        _mark_station_codes(statuses, item["stations_without_fob"], flag="fob_without", category=category, source_row=row_number, label="Without FOB", remarks=item["remarks"])
        _mark_station_codes(statuses, item["stations_with_fob_ramp"], flag="fob_ramp_available", category=category, source_row=row_number, label="FOB/Subway ramp available", remarks=item["remarks"])
        _mark_station_codes(statuses, item["stations_fob_wip"], flag="fob_wip", category=category, source_row=row_number, label="FOB work in progress", remarks=item["remarks"])
        _mark_station_codes(statuses, item["stations_with_lift"], flag="lift_available", category=category, source_row=row_number, label="Lift available", remarks=item["remarks"])
        _mark_station_codes(statuses, item["stations_lift_proposed"], flag="lift_proposed", category=category, source_row=row_number, label="Lift proposed", remarks=item["remarks"])
        _mark_station_codes(statuses, item["stations_ramp_proposed"], flag="ramp_proposed", category=category, source_row=row_number, label="Ramp proposed", remarks=item["remarks"])
        _mark_station_codes(statuses, item["stations_not_feasible_lift_ramp"], flag="not_feasible_lift_ramp", category=category, source_row=row_number, label="Lift/Ramp not feasible", remarks=item["remarks"])

    ramp_feasible_text = _xlsx_cell(rows[28], 1)
    for code in extract_station_codes(ramp_feasible_text):
        row = _status_for(statuses, code)
        row["ramp_feasible"] = True
        _append_status_text(row, "source_rows", "29")
        _append_status_text(row, "status_text", "Station feasible for ramp")

    for row_number in range(40, min(len(rows), 95) + 1):
        row = rows[row_number - 1]
        category_code = _xlsx_cell(row, 4)
        for cell_index in (3, 6):
            for code in extract_station_codes(_xlsx_cell(row, cell_index)):
                item = _status_for(statuses, code)
                if category_code and not item.get("station_detail_category_code"):
                    item["station_detail_category_code"] = category_code
                _append_status_text(item, "source_rows", str(row_number))

    return {"summaries": summaries, "statuses": list(statuses.values())}


def split_scopes(raw: str, station_codes: set[str] | None = None) -> list[dict[str, Any]]:
    text = clean(raw)
    if not text:
        return []
    upper_text = text.upper()
    category_scopes = []
    if re.search(r"\bABSS\b", upper_text):
        category_scopes.append({"scope_type": "ABSS", "scope_value": text, "station_code": None})
    if re.fullmatch(r"DIVISION|DIV|SEC:\s*DIVISION:?|SEC:\s*DIV:?", upper_text) or re.search(r"\bDIVISION\b", upper_text):
        category_scopes.append({"scope_type": "Division", "scope_value": text, "station_code": None})
    body = text
    explicit_station_scope = bool(re.search(r"\bstn:\s*", text, flags=re.I))
    if re.search(r"\bstn:\s*", text, flags=re.I):
        body = re.split(r"\bstn:\s*", text, flags=re.I, maxsplit=1)[1]
    elif re.match(r"^\s*sec\s*:", text, flags=re.I):
        body = re.split(r"^\s*sec\s*:\s*", text, flags=re.I, maxsplit=1)[1]
    body = re.sub(r"\([^)]*\)", " ", body)
    tokens = [token.strip() for token in re.split(r",|;|/|&|\band\b|\n", body, flags=re.I) if token.strip()]
    scopes = [*category_scopes]
    seen = {(scope["scope_type"], scope["scope_value"], scope["station_code"]) for scope in scopes}
    for token in tokens or [body]:
        upper = clean(token).upper()
        if "ABSS" in upper:
            scope = {"scope_type": "ABSS", "scope_value": token, "station_code": None}
            key = (scope["scope_type"], scope["scope_value"], scope["station_code"])
            if key not in seen:
                scopes.append(scope)
                seen.add(key)
        elif re.search(r"\bDIV(ISION)?\b", upper):
            scope = {"scope_type": "Division", "scope_value": token, "station_code": None}
            key = (scope["scope_type"], scope["scope_value"], scope["station_code"])
            if key not in seen:
                scopes.append(scope)
                seen.add(key)
        else:
            candidates = re.findall(r"\b[A-Z]{2,5}[A-Z0-9]?\b", upper)
            matched = candidates if explicit_station_scope else [code for code in candidates if not station_codes or code in station_codes]
            if station_codes and not matched and upper in station_codes:
                matched = [upper]
            for code in matched or ([upper] if not station_codes else []):
                scope = {"scope_type": "Station", "scope_value": code, "station_code": code}
                key = (scope["scope_type"], scope["scope_value"], scope["station_code"])
                if key not in seen:
                    scopes.append(scope)
                    seen.add(key)
    return scopes


def upsert_many(session, model, rows, conflict_cols, update_cols):
    if not rows:
        return 0
    existing_key_cols = [col.key if hasattr(col, "key") else col.name for col in conflict_cols]
    written = 0
    for row in rows:
        filters = {col: row[col] for col in existing_key_cols if row.get(col) is not None}
        obj = session.query(model).filter_by(**filters).one_or_none() if filters else None
        if obj is None:
            obj = model(**row)
            session.add(obj)
        else:
            for col in update_cols:
                setattr(obj, col, row.get(col))
        session.flush()
        written += 1
    return written


def row_to_dict(row) -> dict[str, Any]:
    return {c.name: getattr(row, c.name) for c in row.__table__.columns}


def records_from_query(query) -> list[dict[str, Any]]:
    return [row_to_dict(row) for row in query]


def station_sort_map() -> set[str]:
    return {"station_code", "station_name", "division", "zone", "section", "categorisation", "passenger_footfall"}


def unit_sort_map() -> set[str]:
    return {"unit_no", "station_code", "station_name", "station_category", "licensee_name", "unit_status"}


def work_sort_map() -> set[str]:
    return {"project_id", "source_project_id", "source_sn", "status", "date_of_sanction", "section", "short_name_of_work"}


def earnings_sort_map() -> set[str]:
    return {"receipt_key", "unit_no", "station_code", "date_of_receipt", "licensee_name", "payment_head", "payment_sub_head", "amount", "receipt_type"}


def passenger_amenity_sort_map() -> set[str]:
    return {"station_code", "station_name", "category", "section", "work_type", "tender_status", "sanction_date", "tdc", "platform_count", "available_good_condition", "ramp_feasible", "lift_proposed"}


def commercial_contract_sort_map() -> set[str]:
    return {"contract_name", "licensee_name", "allocation_code", "policy", "sub_category", "asset_scope", "annual_license_fee", "quarterly_license_fee", "contract_period_from", "contract_upto", "station_code", "station_match_status"}


def _canonical_header(value: Any) -> str:
    return normalize(re.sub(r"[^A-Za-z0-9]+", " ", clean(value))).strip()


def _commercial_policy_label(policy: Any) -> str:
    text = clean(policy).strip(" -").upper()
    if text == "ADVERTISING":
        return "Advertising"
    if text == "MSS":
        return "MSS"
    if text == "PARKING":
        return "Parking"
    if "ATM" in text or "BANK" in text or "DBU" in text:
        return "ATM/Banking"
    if "PAY" in text:
        return "Pay and Use"
    if "PMBJK" in text:
        return "PMBJK"
    if "MOBILE" in text:
        return "Mobile Assets"
    return clean(policy).strip(" -") or "Unclassified"


def _commercial_asset_scope(sub_category: Any, station_value: Any) -> str:
    text = normalize(" ".join([clean(sub_category), clean(station_value)]))
    if "out of home" in text or "ooh" in text:
        return "OOH - Out of Home"
    if "parking" in text or "radio taxi" in text or "ev" in text:
        return "Parking / Mobility"
    if "train" in text or "coach" in text or "memu" in text or "interior" in text or "exterior" in text:
        return "Mobile / Train Assets"
    if "atm" in text or "dbu" in text or "bank" in text:
        return "ATM / Banking"
    if "pay and use" in text or "toilet" in text:
        return "Pay and Use"
    if "kiosk" in text or "store" in text or "locker" in text or "pods" in text:
        return "Station Retail / MSS"
    return "Station Commercial"


def _split_station_tokens(value: Any) -> list[str]:
    text = clean(value).upper()
    if not text:
        return []
    return [token.strip() for token in re.split(r"[,;/&]+|\band\b", text, flags=re.I) if token.strip()]


def _infer_station_codes(contract_name: Any, station_codes: set[str]) -> list[str]:
    tokens = re.findall(r"\b[A-Z]{2,5}[A-Z0-9]?\b", clean(contract_name).upper())
    seen = set()
    matches = []
    for token in tokens:
        if token in station_codes and token not in seen:
            seen.add(token)
            matches.append(token)
    return matches


def _commercial_station_links(raw_station: Any, contract_name: Any, station_codes: set[str], asset_scope: str) -> tuple[list[dict[str, Any]], str]:
    raw = clean(raw_station)
    direct_tokens = _split_station_tokens(raw)
    direct_matches = []
    for token in direct_tokens:
        if token in station_codes and token not in direct_matches:
            direct_matches.append(token)
    inferred = [code for code in _infer_station_codes(contract_name, station_codes) if code not in direct_matches]

    links: list[dict[str, Any]] = []
    if direct_matches:
        for code in direct_matches:
            links.append({"station_code": code, "raw_station_value": raw, "match_type": "station_link", "match_status": "Station linked"})
        return links, "Station linked"

    if inferred and asset_scope != "Mobile / Train Assets":
        for code in inferred:
            links.append({"station_code": code, "raw_station_value": raw or clean(contract_name), "match_type": "station_inferred", "match_status": "Station linked"})
        return links, "Station linked"

    if raw or asset_scope == "Mobile / Train Assets":
        links.append({"station_code": None, "raw_station_value": raw or clean(contract_name), "match_type": "asset_scope", "match_status": asset_scope})
        return links, "asset_scope"

    return [], "unmatched"


def parse_commercial_contract_workbook(path_or_bytes: str | bytes) -> dict[str, list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to import commercial contracts") from exc

    source = io.BytesIO(path_or_bytes) if isinstance(path_or_bytes, bytes) else path_or_bytes
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet_name = "exp yearly" if "exp yearly" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"contracts": [], "links": [], "payments": []}

    headers = [cell for cell in rows[0]]
    normalized_headers = [_canonical_header(cell) for cell in headers]
    station_codes = {row.get("station_code") for row in list_stations() if row.get("station_code")}
    station_codes = {clean(code).upper() for code in station_codes}

    month_indexes: list[tuple[int, str, str]] = []
    year_ending_indexes: list[int] = []
    total_2026_indexes: list[int] = []
    for index, header in enumerate(headers):
        parsed = parse_date_value(header)
        if parsed and parsed.day == 1:
            month_indexes.append((index, parsed.isoformat(), str(header)))
        text = _canonical_header(header)
        if text == "year ending":
            year_ending_indexes.append(index)
        if "2026 2027" in text and "total" in text and "license fee" in text:
            total_2026_indexes.append(index)

    def value_for(row: tuple[Any, ...], header_name: str) -> Any:
        wanted = _canonical_header(header_name)
        for index, header in enumerate(normalized_headers):
            if header == wanted:
                return row[index] if index < len(row) else None
        return None

    contracts: list[dict[str, Any]] = []
    links: list[dict[str, Any]] = []
    payments: list[dict[str, Any]] = []
    for row in rows[1:]:
        contract_name = clean(value_for(row, "Contract Name"))
        if not contract_name:
            continue

        raw_station = clean(value_for(row, "Stn"))
        policy = _commercial_policy_label(value_for(row, "POLICY"))
        sub_category = clean(value_for(row, "SUB-CATEGORY"))
        asset_scope = _commercial_asset_scope(sub_category, raw_station)
        station_links, station_match_status = _commercial_station_links(raw_station, contract_name, station_codes, asset_scope)

        contract = {
            "source_sl_no": to_int(value_for(row, "sl no")),
            "raw_station_value": raw_station,
            "contract_name": contract_name,
            "licensee_name": clean(value_for(row, "LICENSEE")),
            "allocation_code": clean(value_for(row, "ALLOCATION CODE")),
            "contract_allotted_on": date_text(value_for(row, "CONTRACT ALLOTTED ON")),
            "policy": policy,
            "sub_category": sub_category,
            "asset_scope": asset_scope,
            "space_sq_ft": to_int(value_for(row, "Space Allotted Sq Ft")),
            "annual_license_fee": to_int(value_for(row, "Annual License fee")),
            "quarterly_license_fee": to_int(value_for(row, "QUATERLY LICENSE FEE")),
            "security_deposit": to_int(value_for(row, "Security Deposit")),
            "no_of_years": to_int(value_for(row, "No of years")),
            "contract_period_from": date_text(value_for(row, "Contract period From")),
            "contract_upto": date_text(value_for(row, "contract Upto")),
            "cycle": clean(value_for(row, "CYCLE")),
            "year_ending_amount": to_int(row[year_ending_indexes[0]]) if year_ending_indexes and year_ending_indexes[0] < len(row) else None,
            "total_license_fee_2026_2027": to_int(row[total_2026_indexes[0]]) if total_2026_indexes and total_2026_indexes[0] < len(row) else None,
            "renewal_status": clean(value_for(row, "Renewal Status")),
            "termination_status": clean(value_for(row, "Termination Status")),
            "tender_status": clean(value_for(row, "Tender Status")),
            "station_match_status": station_match_status,
        }
        contracts.append(contract)
        for link in station_links:
            links.append({"contract_name": contract_name, **link})
        for index, month, source_column in month_indexes:
            amount = to_int(row[index]) if index < len(row) else None
            if amount is None:
                continue
            payments.append({
                "contract_name": contract_name,
                "payment_month": month,
                "source_column": source_column,
                "amount_due": amount,
                "amount_paid": amount,
                "payment_status": "Recorded",
            })

    return {"contracts": contracts, "links": links, "payments": payments}


def list_passenger_amenities(kind: str = "summary", q: str | None = None, station_code: str | None = None) -> list[dict[str, Any]]:
    session = SessionLocal()
    try:
        like = f"%{q}%" if q else None
        if kind == "norms":
            query = session.query(AmenityNorm)
            if like:
                query = query.filter((AmenityNorm.category.ilike(like)) | (AmenityNorm.amenity.ilike(like)) | (AmenityNorm.norm.ilike(like)))
            return [row_to_dict(row) for row in query.order_by(AmenityNorm.category, AmenityNorm.amenity, AmenityNorm.norm).all()]

        station_rows = session.query(Station).filter(
            Station.is_active.is_(True),
            func.lower(func.trim(Station.categorisation)).notin_(("", "test", "non-commercial")),
        ).order_by(Station.station_name, Station.station_code).all()
        stations_by_code = {station.station_code: station for station in station_rows}
        source_code_aliases = {"GNBH": "GNB"}
        legacy_code_by_canonical = {canonical: source for source, canonical in source_code_aliases.items()}

        def canonical_station(source_code: Any) -> Station | None:
            code = clean(source_code).upper()
            return stations_by_code.get(source_code_aliases.get(code, code))

        def amenity_row(source_row: Any) -> dict[str, Any] | None:
            row = row_to_dict(source_row)
            station = canonical_station(row.get("station_code"))
            if not station:
                return None
            row.update({
                "station_code": station.station_code,
                "station_name": station.station_name,
                "division": station.division,
                "section": station.section,
                "category": station.categorisation,
                "categorisation": station.categorisation,
            })
            return row

        selected_station = clean(station_code).upper() if station_code and station_code != "All" else ""

        def row_matches(row: dict[str, Any]) -> bool:
            if selected_station and normalize(row.get("station_code")) != normalize(selected_station):
                return False
            return not q or any(normalize(q) in normalize(value) for value in row.values())

        def canonical_rows(source_rows: list[Any]) -> list[dict[str, Any]]:
            rows = []
            for source_row in source_rows:
                row = amenity_row(source_row)
                if row and row_matches(row):
                    rows.append(row)
            return rows

        if kind == "infra":
            infra_by_station = {row.station_code: row for row in session.query(StationInfra).all()}
            rows = []
            for station in station_rows:
                if selected_station and normalize(station.station_code) != normalize(selected_station):
                    continue
                legacy_source_code = legacy_code_by_canonical.get(station.station_code)
                infra = infra_by_station.get(station.station_code) or infra_by_station.get(legacy_source_code)
                row = row_to_dict(infra) if infra else {
                    "infra_key": None,
                    "platform_list": None,
                    "platform_count": None,
                    "platform_level": None,
                    "fob_details": None,
                    "shelter_details": None,
                    "remarks": None,
                }
                row.update({
                    "station_code": station.station_code,
                    "station_name": station.station_name,
                    "category": station.categorisation,
                    "categorisation": station.categorisation,
                    "division": station.division,
                    "section": station.section,
                })
                if row_matches(row):
                    rows.append(row)
            return rows
        if kind == "platforms":
            return canonical_rows(session.query(PlatformDetail).order_by(PlatformDetail.station_code, PlatformDetail.platform).all())
        if kind == "wheelchairs":
            return canonical_rows(session.query(WheelChairAvailability).order_by(WheelChairAvailability.station_code).all())
        if kind == "trolley":
            return canonical_rows(session.query(TrolleyPath).order_by(TrolleyPath.station_code).all())
        if kind == "pa_works":
            return canonical_rows(session.query(PassengerAmenityWork).order_by(PassengerAmenityWork.work_type, PassengerAmenityWork.station_code).all())
        if kind == "pf_extension":
            return canonical_rows(session.query(StationPlatformExtensionStatus).order_by(StationPlatformExtensionStatus.station_code).all())
        if kind == "pf_extension_summary":
            query = session.query(PlatformExtensionSummary)
            if like:
                query = query.filter((PlatformExtensionSummary.category.ilike(like)) | (PlatformExtensionSummary.remarks.ilike(like)) | (PlatformExtensionSummary.summary_type.ilike(like)))
            return [row_to_dict(row) for row in query.order_by(PlatformExtensionSummary.summary_type, PlatformExtensionSummary.category).all()]

        infra_by_station = {row.station_code: row for row in session.query(StationInfra).all()}
        wheel_by_station = {row.station_code: row for row in session.query(WheelChairAvailability).all()}
        trolley_by_station = {row.station_code: row for row in session.query(TrolleyPath).all()}
        platform_counts: dict[str, int] = {}
        for source_code, count in session.execute(select(PlatformDetail.station_code, func.count(PlatformDetail.platform_key)).group_by(PlatformDetail.station_code)).all():
            canonical = canonical_station(source_code)
            if canonical:
                platform_counts[canonical.station_code] = platform_counts.get(canonical.station_code, 0) + count
        pa_work_counts = dict(session.execute(select(PassengerAmenityWork.station_code, func.count(PassengerAmenityWork.pa_work_key)).group_by(PassengerAmenityWork.station_code)).all())
        pf_status_by_station = {row.station_code: row for row in session.query(StationPlatformExtensionStatus).all()}
        rows = []
        for station in station_rows:
            if station_code and station_code != "All" and station.station_code != station_code:
                continue
            infra = infra_by_station.get(station.station_code) or infra_by_station.get(legacy_code_by_canonical.get(station.station_code))
            wheel = wheel_by_station.get(station.station_code)
            trolley = trolley_by_station.get(station.station_code)
            row = {
                "station_code": station.station_code,
                "station_name": station.station_name,
                "division": station.division,
                "section": station.section,
                "category": station.categorisation,
                "platform_type": station.platform_type,
                "platform_count": infra.platform_count if infra else station.number_of_platforms,
                "platform_detail_count": platform_counts.get(station.station_code, 0),
                "wheel_chairs": wheel.available_good_condition if wheel else None,
                "trolley_path": trolley.trolley_path if trolley else None,
                "pa_works": pa_work_counts.get(station.station_code, 0),
                "fob_details": infra.fob_details if infra else None,
                "ramp_feasible": pf_status_by_station.get(station.station_code).ramp_feasible if pf_status_by_station.get(station.station_code) else False,
                "lift_proposed": pf_status_by_station.get(station.station_code).lift_proposed if pf_status_by_station.get(station.station_code) else False,
                "pf_extension_proposed": pf_status_by_station.get(station.station_code).pf_extension_proposed if pf_status_by_station.get(station.station_code) else False,
            }
            if not q or any(normalize(q) in normalize(value) for value in row.values()):
                rows.append(row)
        return sorted(rows, key=lambda item: (item.get("station_name") or "", item.get("station_code") or ""))
    finally:
        session.close()


def get_passenger_amenity_reports(amenity_data: dict[str, list[dict[str, Any]]] | None = None) -> dict[str, Any]:
    # Build every KPI from the same canonical rows returned to the UI. Source
    # sheets contain totals and section headings that must never count as
    # stations or make the KPI disagree with a filtered table.
    data = amenity_data or {
        "summary": list_passenger_amenities(kind="summary"),
        "infra": list_passenger_amenities(kind="infra"),
        "platforms": list_passenger_amenities(kind="platforms"),
        "wheelchairs": list_passenger_amenities(kind="wheelchairs"),
        "trolley": list_passenger_amenities(kind="trolley"),
        "works": list_passenger_amenities(kind="pa_works"),
        "norms": list_passenger_amenities(kind="norms"),
        "pfExtension": list_passenger_amenities(kind="pf_extension"),
    }
    stations = data.get("summary") or []
    infra = data.get("infra") or []
    platforms = data.get("platforms") or []
    wheelchairs = data.get("wheelchairs") or []
    trolley = data.get("trolley") or []
    pa_work_rows = data.get("works") or []
    norms_rows = data.get("norms") or []
    pf_rows = data.get("pfExtension") or []

    station_count = len(stations)
    infra_count = len(infra)
    wheelchair_stations = len({row.get("station_code") for row in wheelchairs if row.get("station_code")})
    pf_statuses = len({row.get("station_code") for row in pf_rows if row.get("station_code")})
    return {
        "stations": station_count,
        "infra_records": infra_count,
        "platform_records": len(platforms),
        "wheelchair_stations": wheelchair_stations,
        "trolley_path_yes": sum("yes" in normalize(row.get("trolley_path")) for row in trolley),
        "trolley_path_no": sum("no" in normalize(row.get("trolley_path")) for row in trolley),
        "pa_works": len(pa_work_rows),
        "open_pa_works": sum("complete" not in normalize(row.get("progress")) for row in pa_work_rows),
        "norms": len(norms_rows),
        "pf_extension_statuses": pf_statuses,
        "pf_extension_wip": sum(bool(row.get("pf_extension_wip")) for row in pf_rows),
        "pf_extension_proposed": sum(bool(row.get("pf_extension_proposed")) for row in pf_rows),
        "ramp_feasible": sum(bool(row.get("ramp_feasible")) for row in pf_rows),
        "lift_proposed": sum(bool(row.get("lift_proposed")) for row in pf_rows),
        "ramp_proposed": sum(bool(row.get("ramp_proposed")) for row in pf_rows),
        "not_feasible_lift_ramp": sum(bool(row.get("not_feasible_lift_ramp")) for row in pf_rows),
        "coverage": {
            "infra": round((infra_count / station_count) * 100, 1) if station_count else 0,
            "wheelchairs": round((wheelchair_stations / station_count) * 100, 1) if station_count else 0,
            "pf_extension": round((pf_statuses / station_count) * 100, 1) if station_count else 0,
        },
    }


def get_stats() -> dict[str, Any]:
    session = SessionLocal()
    try:
        contract_earnings = session.query(Earning).filter(
            (Earning.earning_scope.is_(None)) | (Earning.earning_scope != "tender_emd")
        )
        commercial_station_filter = (
            Station.is_active.is_(True),
            func.lower(func.trim(Station.categorisation)).notin_(("", "test", "non-commercial")),
        )
        return {
            "stations": session.query(func.count(Station.station_code)).filter(*commercial_station_filter).scalar() or 0,
            "units": session.query(func.count(Unit.unit_no)).scalar() or 0,
            "works": session.query(func.count(Work.work_key)).scalar() or 0,
            "earnings": contract_earnings.count(),
            "commercialContracts": session.query(func.count(CommercialContract.contract_key)).scalar() or 0,
            "commercialRevenue": session.query(func.coalesce(func.sum(CommercialContract.annual_license_fee), 0)).scalar() or 0,
            "links": session.query(func.count(WorkLink.id)).scalar() or 0,
            "earningsTotal": contract_earnings.with_entities(func.coalesce(func.sum(Earning.amount), 0)).scalar() or 0,
            "footfall": session.query(func.coalesce(func.sum(Station.passenger_footfall), 0)).filter(*commercial_station_filter).scalar() or 0,
            "topStations": [
                {"station_code": code, "station_name": name, "works": works}
                for code, name, works in session.execute(
                    select(Station.station_code, Station.station_name, func.count(WorkLink.id).label("works"))
                    .join(WorkLink, WorkLink.station_code == Station.station_code, isouter=True)
                    .where(*commercial_station_filter)
                    .group_by(Station.station_code, Station.station_name, Station.passenger_footfall)
                    .order_by(func.count(WorkLink.id).desc(), Station.passenger_footfall.desc().nullslast())
                    .limit(8)
                )
            ],
        }
    finally:
        session.close()


def get_data_centre_status() -> dict[str, Any]:
    """Return one freshness and data-quality view for every managed module."""
    session = SessionLocal()
    try:
        commercial_station_filter = (
            Station.is_active.is_(True),
            func.lower(func.trim(Station.categorisation)).notin_(('', 'test', 'non-commercial')),
        )

        module_models = {
            "stations": Station,
            "units": Unit,
            "earnings": Earning,
            "works": Work,
            "work_progress": WorkProgressUpdate,
            "contracts": CommercialContract,
        }
        modules = {}
        for name, model in module_models.items():
            query = session.query(model)
            if name == "stations":
                query = query.filter(*commercial_station_filter)
            last_updated_query = session.query(func.max(model.updated_at))
            if hasattr(model, "is_active"):
                last_updated_query = last_updated_query.filter(model.is_active.is_(True))
            last_updated = last_updated_query.scalar()
            modules[name] = {
                "count": query.count(),
                "last_updated_at": last_updated.isoformat() if last_updated else None,
            }

        amenity_counts = {
            "norms": session.query(AmenityNorm).filter(AmenityNorm.is_active.is_(True)).count(),
            "infra": session.query(StationInfra).filter(StationInfra.is_active.is_(True)).count(),
            "platforms": session.query(PlatformDetail).filter(PlatformDetail.is_active.is_(True)).count(),
            "wheelchairs": session.query(WheelChairAvailability).filter(WheelChairAvailability.is_active.is_(True)).count(),
            "trolley": session.query(TrolleyPath).filter(TrolleyPath.is_active.is_(True)).count(),
            "works": session.query(PassengerAmenityWork).filter(PassengerAmenityWork.is_active.is_(True)).count(),
            "accessibility": session.query(StationPlatformExtensionStatus).filter(StationPlatformExtensionStatus.is_active.is_(True)).count(),
        }
        amenity_models = (
            AmenityNorm,
            StationInfra,
            PlatformDetail,
            WheelChairAvailability,
            TrolleyPath,
            PassengerAmenityWork,
            StationPlatformExtensionStatus,
        )
        amenity_dates = [
            session.query(func.max(model.updated_at))
            .filter(model.is_active.is_(True))
            .scalar()
            for model in amenity_models
        ]
        amenity_last_updated = max((value for value in amenity_dates if value), default=None)
        modules["amenities"] = {
            "count": sum(amenity_counts.values()),
            "last_updated_at": amenity_last_updated.isoformat() if amenity_last_updated else None,
            "breakdown": amenity_counts,
        }

        station_codes = {row[0] for row in session.query(Station.station_code).all()}
        unit_codes = {row[0] for row in session.query(Unit.unit_no).all()}
        quality = {
            "units_missing_station": sum(1 for row in session.query(Unit.station_code).all() if not row[0] or row[0] not in station_codes),
            "earnings_missing_unit": sum(1 for row in session.query(Earning.unit_no).all() if not row[0] or row[0] not in unit_codes),
            "earnings_missing_station": sum(1 for row in session.query(Earning.station_code).all() if not row[0] or row[0] not in station_codes),
            "works_unmatched_scope": session.query(WorkLink).filter(WorkLink.match_status != "Matched").count(),
            "contracts_unmatched_station": session.query(CommercialContractStationLink).filter(CommercialContractStationLink.match_status != "Matched").count(),
        }
        quality["total"] = sum(quality.values())
        quality["exceptions"] = []
        for row in (
            session.query(Unit.unit_no, Unit.station_code)
            .filter((Unit.station_code.is_(None)) | (~Unit.station_code.in_(station_codes)))
            .limit(50)
            .all()
        ):
            quality["exceptions"].append({
                "module": "units",
                "record_key": row.unit_no,
                "station_code": row.station_code,
                "problem": "Missing or unmatched station link",
            })
        for row in (
            session.query(Earning.earning_key, Earning.unit_no, Earning.station_code)
            .filter(
                (Earning.unit_no.is_(None))
                | (~Earning.unit_no.in_(unit_codes))
                | (Earning.station_code.is_(None))
                | (~Earning.station_code.in_(station_codes))
            )
            .limit(50)
            .all()
        ):
            quality["exceptions"].append({
                "module": "earnings",
                "record_key": row.earning_key,
                "station_code": row.station_code,
                "problem": "Missing or unmatched unit/station link",
            })
        for row in (
            session.query(WorkLink.project_id, WorkLink.station_code, WorkLink.match_status)
            .filter(WorkLink.match_status != "Matched")
            .limit(50)
            .all()
        ):
            quality["exceptions"].append({
                "module": "works",
                "record_key": row.project_id,
                "station_code": row.station_code,
                "problem": row.match_status or "Unmatched work scope",
            })
        for row in (
            session.query(CommercialContractStationLink.contract_key, CommercialContractStationLink.station_code, CommercialContractStationLink.match_status)
            .filter(CommercialContractStationLink.match_status != "Matched")
            .limit(50)
            .all()
        ):
            quality["exceptions"].append({
                "module": "contracts",
                "record_key": row.contract_key,
                "station_code": row.station_code,
                "problem": row.match_status or "Unmatched contract scope",
            })
        quality["exceptions"] = quality["exceptions"][:150]

        latest_failed_catering = (
            session.query(CateringSyncRun)
            .filter(CateringSyncRun.status == "failed")
            .order_by(CateringSyncRun.started_at.desc())
            .first()
        )
        failures = []
        if latest_failed_catering:
            failures.append({
                "resource": "catering",
                "message": latest_failed_catering.error_message or "Catering synchronization failed",
                "at": latest_failed_catering.started_at.isoformat() if latest_failed_catering.started_at else None,
            })

        changes = (
            session.query(DataChangeLog)
            .order_by(DataChangeLog.created_at.desc())
            .limit(20)
            .all()
        )
        for row in changes:
            action = normalize(row.action)
            if any(token in action for token in ("fail", "error", "rollback")):
                failures.append({
                    "resource": row.resource,
                    "message": row.details or row.action,
                    "at": row.created_at.isoformat() if row.created_at else None,
                })
        failures = failures[:50]
        recent_sync = [
            {
                "resource": row.resource,
                "action": row.action,
                "source": row.source,
                "details": row.details,
                "at": row.created_at.isoformat() if row.created_at else None,
            }
            for row in changes
        ]
        latest_successful_sync = recent_sync[0]["at"] if recent_sync else None
        source_snapshots = {
            name: {
                "source": None,
                "source_count": None,
                "last_refresh_at": None,
                "details": None,
            }
            for name in modules
        }
        for row in changes:
            resource = {
                "catering": "catering",
                "commercial_contracts": "contracts",
                "passenger_amenities": "amenities",
            }.get(row.resource, row.resource)
            targets = ["units", "earnings"] if resource == "catering" else [resource]
            for target in targets:
                if target not in source_snapshots or source_snapshots[target]["last_refresh_at"]:
                    continue
                details = row.details or ""
                match = re.search(r"(\d+)\s+rows", details)
                source_snapshots[target] = {
                    "source": row.source,
                    "source_count": int(match.group(1)) if match else None,
                    "last_refresh_at": row.created_at.isoformat() if row.created_at else None,
                    "details": details,
                }
        catering_run = (
            session.query(CateringSyncRun)
            .filter(CateringSyncRun.status == "success")
            .order_by(CateringSyncRun.started_at.desc())
            .first()
        )
        if catering_run:
            reconciliation = {}
            if catering_run.report_json:
                try:
                    report_json = json.loads(catering_run.report_json)
                    applied = report_json.get("applied", {})
                    reconciliation = {
                        "added": applied.get("inserted_units", 0) + applied.get("inserted_earnings", 0),
                        "changed": applied.get("updated_units", 0) + applied.get("updated_earnings", 0),
                        "removed": applied.get("deactivated_units", 0) + applied.get("removed_stale_earnings", 0),
                        "unmatched": report_json.get("reconciliation", {}).get("unlinked_earning_rows", 0),
                    }
                except (TypeError, ValueError):
                    reconciliation = {}
            for target, count in (("units", catering_run.unit_rows), ("earnings", catering_run.earning_rows)):
                source_snapshots[target] = {
                    "source": "google_sheet",
                    "source_count": count,
                    "last_refresh_at": catering_run.completed_at.isoformat() if catering_run.completed_at else catering_run.started_at.isoformat(),
                    "details": "Catering sync history",
                    "reconciliation": reconciliation,
                }
        # Always expose a useful baseline in the Data Centre, even for modules
        # whose last import predates the current change log retention window.
        for target, module in modules.items():
            snapshot = source_snapshots[target]
            if snapshot["source_count"] is None:
                snapshot.update({
                    "source": "postgresql",
                    "source_count": module["count"],
                    "last_refresh_at": module["last_updated_at"],
                    "details": "Current PostgreSQL baseline; source history unavailable",
                })
        device_states = session.query(MobileDeviceState).order_by(MobileDeviceState.last_seen_at.desc()).all()
        mobile_cache = {
            "device_count": len(device_states),
            "active_device_count": sum(
                1
                for device in device_states
                if device.last_seen_at and device.last_seen_at >= datetime.now(timezone.utc) - timedelta(days=7)
            ),
            "last_seen_at": device_states[0].last_seen_at.isoformat() if device_states and device_states[0].last_seen_at else None,
            "data_version": next((device.data_version for device in device_states if device.data_version), None),
            "counts": {
                "stations": sum(device.cached_stations for device in device_states),
                "station_details": sum(device.cached_station_details for device in device_states),
                "works": sum(device.cached_works for device in device_states),
                "units": sum(device.cached_units for device in device_states),
                "earnings": sum(device.cached_earnings for device in device_states),
            },
            "pending_operations": sum(device.pending_operations for device in device_states),
            "failed_operations": sum(device.failed_operations for device in device_states),
        }
        return {
            "status": "attention" if failures or quality["total"] else "ready",
            "modules": modules,
            "quality": quality,
            "failures": failures,
            "recent_sync": recent_sync,
            "source_snapshots": source_snapshots,
            "mobile_cache": mobile_cache,
            "last_sync_at": latest_successful_sync,
        }
    finally:
        session.close()


def list_stations(q: str | None = None, category: str | None = None) -> list[dict[str, Any]]:
    session = SessionLocal()
    try:
        query = session.query(Station).filter(
            Station.is_active.is_(True),
            func.lower(func.trim(Station.categorisation)).notin_(('', 'test', 'non-commercial')),
        )
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Station.station_code.ilike(like))
                | (Station.station_name.ilike(like))
                | (Station.section.ilike(like))
                | (Station.division.ilike(like))
                | (Station.categorisation.ilike(like))
            )
        if category and category != "All":
            query = query.filter(Station.categorisation == category)
        return [row_to_dict(row) for row in query.order_by(Station.station_name, Station.station_code).all()]
    finally:
        session.close()


def list_units(q: str | None = None, station_code: str | None = None) -> list[dict[str, Any]]:
    session = SessionLocal()
    try:
        query = session.query(Unit, Station.station_name, Station.categorisation).join(Station, Station.station_code == Unit.station_code, isouter=True)
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Unit.unit_no.ilike(like))
                | (Unit.licensee_name.ilike(like))
                | (Unit.station_code.ilike(like))
                | (Unit.unit_status.ilike(like))
            )
        if station_code and station_code != "All":
            query = query.filter(Unit.station_code == station_code)
        raw_rows = query.order_by(Unit.station_code, Unit.unit_no).all()
        unit_numbers = [unit.unit_no for unit, _, _ in raw_rows if unit.unit_no]
        pending_by_unit: dict[str, int] = {}
        if unit_numbers:
            earning_rows = session.query(Earning).filter(Earning.unit_no.in_(unit_numbers)).all()
            for earning in earning_rows:
                receipt_text = normalize(" ".join([clean(earning.receipt_type), clean(earning.payment_head), clean(earning.payment_sub_head)]))
                if "pending" in receipt_text or "due" in receipt_text or "outstanding" in receipt_text:
                    pending_by_unit[earning.unit_no] = pending_by_unit.get(earning.unit_no, 0) + to_money(earning.amount)
        rows = []
        for unit, station_name, categorisation in raw_rows:
            row = {**row_to_dict(unit), "station_name": station_name, "categorisation": categorisation}
            available = is_available_unit(row)
            paid_upto = parse_date_value(unit.paid_upto)
            license_fee = to_money(unit.license_fee)
            if available or not license_fee:
                months_pending = 0
            elif paid_upto is None:
                months_pending = 1
            elif paid_upto < date.today():
                months_pending = month_delta(paid_upto + timedelta(days=1), month_end(date.today()))
            else:
                months_pending = 0
            estimated_pending = months_pending * license_fee
            pending_amount = max(pending_by_unit.get(unit.unit_no, 0), estimated_pending)
            row.update(contract_risk(
                row.get("contract_to"),
                pending_amount=pending_amount,
                station_match_status="Station linked" if unit.station_code else "Missing link",
                missing_fee=not available and not clean(row.get("license_fee")),
            ))
            row["months_pending"] = months_pending
            row["estimated_pending_amount"] = estimated_pending
            row["pending_receipts"] = 1 if pending_amount else 0
            rows.append(row)
        return rows
    finally:
        session.close()


def list_works(q: str | None = None, scope_type: str | None = None, station_code: str | None = None) -> list[dict[str, Any]]:
    session = SessionLocal()
    try:
        query = session.query(
            Work,
            WorkLink,
            Station.station_name,
            Station.categorisation,
            Station.sr_den,
            Station.cmi,
        ).join(WorkLink, WorkLink.project_id == Work.project_id, isouter=True).join(Station, Station.station_code == WorkLink.station_code, isouter=True)
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Work.project_id.ilike(like))
                | (Work.short_name_of_work.ilike(like))
                | (Work.block_section_station.ilike(like))
                | (WorkLink.scope_value.ilike(like))
                | (WorkLink.scope_type.ilike(like))
            )
        if scope_type and scope_type != "All":
            query = query.filter(WorkLink.scope_type == scope_type)
        if station_code and station_code != "All":
            query = query.filter(WorkLink.station_code == station_code)
        grouped: dict[str, dict[str, Any]] = {}
        for work, wl, station_name, categorisation, sr_den, cmi in query.order_by(Work.date_of_sanction.desc().nullslast(), Work.project_id).all():
            key = work.project_id
            row = grouped.get(key)
            if row is None:
                row = {
                    **row_to_dict(work),
                    "scope_type": wl.scope_type if wl else "Unlinked",
                    "scope_value": wl.scope_value if wl else work.block_section_station,
                    "station_code": wl.station_code if wl else None,
                    "match_status": wl.match_status if wl else "Missing link",
                    "station_name": station_name,
                    "categorisation": categorisation,
                    "sr_den": sr_den,
                    "cmi": cmi,
                    "station_codes": [],
                    "station_names": [],
                    "links": [],
                }
                grouped[key] = row

            if wl and wl.station_code:
                if wl.station_code not in row["station_codes"]:
                    row["station_codes"].append(wl.station_code)
                if station_name and station_name not in row["station_names"]:
                    row["station_names"].append(station_name)
                row["links"].append({
                    "station_code": wl.station_code,
                    "station_name": station_name,
                    "scope_type": wl.scope_type,
                    "scope_value": wl.scope_value,
                    "match_status": wl.match_status,
                })

        rows = list(grouped.values())
        for row in rows:
            row["station_code"] = row["station_codes"][0] if row["station_codes"] else row["station_code"]
            row["station_name"] = row["station_names"][0] if row["station_names"] else row["station_name"]
            row["linked_station_count"] = len(row["station_codes"])
        return rows
    finally:
        session.close()


def get_work_monitoring(
    q: str | None = None,
    section: str | None = None,
    allocation: str | None = None,
    year: str | None = None,
    work_type: str | None = None,
    station_code: str | None = None,
    today: date | None = None,
) -> dict[str, Any]:
    """Build an exception-first monitoring projection from master and history tables."""
    today = today or date.today()
    session = SessionLocal()
    try:
        query = (
            session.query(Work, WorkLink, Station.station_name, Station.sr_den, Station.cmi)
            .join(WorkLink, WorkLink.project_id == Work.project_id, isouter=True)
            .join(Station, Station.station_code == WorkLink.station_code, isouter=True)
        )
        if section and section != "All":
            query = query.filter(Work.section == section)
        if allocation and allocation != "All":
            query = query.filter(Work.allocation == allocation)
        if year and year != "All":
            query = query.filter(Work.year_of_sanction == year)
        if station_code and station_code != "All":
            query = query.filter(WorkLink.station_code == station_code.upper())

        latest_progress: dict[str, dict[str, Any]] = {}
        progress_count: dict[str, int] = {}
        progress_rows = session.query(WorkProgressUpdate).order_by(WorkProgressUpdate.update_date.desc(), WorkProgressUpdate.progress_id.desc()).all()
        for row in progress_rows:
            progress_count[row.project_id] = progress_count.get(row.project_id, 0) + 1
            latest_progress.setdefault(row.project_id, row_to_dict(row))
        latest_expenditure: dict[str, dict[str, Any]] = {}
        expenditure_count: dict[str, int] = {}
        expenditure_rows = session.query(WorkExpenditureUpdate).order_by(WorkExpenditureUpdate.update_date.desc(), WorkExpenditureUpdate.expenditure_id.desc()).all()
        for row in expenditure_rows:
            expenditure_count[row.project_id] = expenditure_count.get(row.project_id, 0) + 1
            latest_expenditure.setdefault(row.project_id, row_to_dict(row))

        def progress_percent(value: Any) -> int | None:
            match = re.search(r"(\d+(?:\.\d+)?)", clean(value))
            return round(float(match.group(1))) if match else None

        def classify(work: Work, link: WorkLink | None) -> str:
            text = normalize(" ".join([clean(work.short_name_of_work), clean(work.remarks), clean(work.section), clean(link.scope_type if link else "")]))
            if "abss" in text:
                return "ABSS"
            if "cao/cn" in text or "cao cn" in text:
                return "CAO/CN"
            if "fob" in text or "foot over" in text:
                return "FOB"
            if "platform" in text or "raising" in text:
                return "Platform"
            if "divyang" in text or "ramp" in text or "lift" in text:
                return "Divyangjan"
            if "goods" in text or "csgr" in text:
                return "Goods / CSGR"
            return clean(link.scope_type if link else "Other") or "Other"

        grouped: dict[str, dict[str, Any]] = {}
        for work, link, station_name, sr_den, cmi in query.all():
            item = grouped.get(work.project_id)
            if not item:
                latest = latest_progress.get(work.project_id) or {}
                expenditure = latest_expenditure.get(work.project_id) or {}
                status = clean(latest.get("status") or work.status) or "Unknown"
                percentage = latest.get("progress_percent")
                if percentage is None:
                    percentage = progress_percent(latest.get("physical_progress") or work.physical_progress)
                completed = bool(re.search(r"complete|done", status, flags=re.I))
                if percentage is not None and percentage >= 100:
                    completed = True
                tdc = latest.get("tdc") or work.tdc
                tdc_date = parse_date_value(tdc)
                days_to_tdc = (tdc_date - today).days if tdc_date else None
                contradiction = None
                if completed and percentage is not None and percentage < 100:
                    contradiction = "Completed status with progress below 100%"
                elif not completed and percentage is not None and percentage >= 100:
                    contradiction = "100% progress with an open status"
                if contradiction:
                    alert_type = "contradiction"
                elif not completed and days_to_tdc is not None and days_to_tdc < 0:
                    alert_type = "tdc_overdue"
                elif not completed and days_to_tdc is not None and days_to_tdc <= 30:
                    alert_type = "tdc_due_30_days"
                else:
                    alert_type = None
                item = {
                    **row_to_dict(work),
                    "station_code": link.station_code if link else None,
                    "station_name": station_name,
                    "sr_den": sr_den,
                    "cmi": cmi,
                    "work_type": classify(work, link),
                    "effective_status": status,
                    "progress_percent": percentage,
                    "completed": completed,
                    "tdc": tdc,
                    "days_to_tdc": days_to_tdc,
                    "alert_type": alert_type,
                    "alert_label": "TDC overdue" if alert_type == "tdc_overdue" else "TDC due within 30 days" if alert_type == "tdc_due_30_days" else contradiction,
                    "latest_progress_update": latest.get("update_date"),
                    "progress_update_count": progress_count.get(work.project_id, 0),
                    "latest_expenditure_update": expenditure.get("update_date"),
                    "expenditure_update_count": expenditure_count.get(work.project_id, 0),
                    "latest_cumulative_expenditure": expenditure.get("cumulative_expenditure") if expenditure else work.expenditure_upto_date,
                    "station_codes": [],
                }
                grouped[work.project_id] = item
            if link and link.station_code and link.station_code not in item["station_codes"]:
                item["station_codes"].append(link.station_code)

        items = list(grouped.values())
        if q:
            query_text = normalize(q)
            items = [item for item in items if query_text in normalize(" ".join(str(item.get(key) or "") for key in ("project_id", "short_name_of_work", "section", "allocation", "work_type", "station_name", "station_code", "sr_den", "cmi")))]
        if work_type and work_type != "All":
            items = [item for item in items if item["work_type"] == work_type]
        items.sort(key=lambda item: (item["alert_type"] is None, item["days_to_tdc"] is None, item["days_to_tdc"] if item["days_to_tdc"] is not None else 99999, item["project_id"] or ""))

        def counts(key: str) -> list[dict[str, Any]]:
            result: dict[str, int] = {}
            for item in items:
                label = clean(item.get(key)) or "Unknown"
                result[label] = result.get(label, 0) + 1
            return [{"label": label, "value": value} for label, value in sorted(result.items(), key=lambda pair: (-pair[1], pair[0]))]

        return {
            "as_of": today.isoformat(),
            "total": len(items),
            "open": sum(1 for item in items if not item["completed"]),
            "completed": sum(1 for item in items if item["completed"]),
            "exceptions": sum(1 for item in items if item["alert_type"]),
            "tdc_overdue": sum(1 for item in items if item["alert_type"] == "tdc_overdue"),
            "tdc_due_30_days": sum(1 for item in items if item["alert_type"] == "tdc_due_30_days"),
            "contradictions": sum(1 for item in items if item["alert_type"] == "contradiction"),
            "by_section": counts("section"),
            "by_allocation": counts("allocation"),
            "by_year": counts("year_of_sanction"),
            "by_work_type": counts("work_type"),
            "items": items,
        }
    finally:
        session.close()


def list_earnings(
    q: str | None = None,
    unit_no: str | None = None,
    station_code: str | None = None,
    include_tender_emd: bool = False,
) -> list[dict[str, Any]]:
    session = SessionLocal()
    try:
        query = session.query(Earning, EarningLink, Unit.station_category, Unit.unit_status, Unit.type_of_unit, Station.station_name, Station.division, Station.section, Station.categorisation).join(EarningLink, EarningLink.receipt_key == Earning.receipt_key, isouter=True).join(Unit, Unit.unit_no == Earning.unit_no, isouter=True).join(Station, Station.station_code == func.coalesce(Earning.station_code, Unit.station_code), isouter=True)
        if not include_tender_emd:
            query = query.filter((Earning.earning_scope.is_(None)) | (Earning.earning_scope != "tender_emd"))
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Earning.unit_no.ilike(like))
                | (Earning.station_code.ilike(like))
                | (Earning.licensee_name.ilike(like))
                | (Earning.payment_head.ilike(like))
                | (Earning.payment_sub_head.ilike(like))
                | (Earning.mr_no.ilike(like))
            )
        if unit_no and unit_no != "All":
            query = query.filter(Earning.unit_no == unit_no)
        if station_code and station_code != "All":
            query = query.filter(Earning.station_code == station_code)
        return [{**row_to_dict(earning), "match_status": link.match_status if link else None, "station_category": station_category, "unit_status": unit_status, "type_of_unit": type_of_unit, "station_name": station_name, "division": division, "section": section, "categorisation": categorisation} for earning, link, station_category, unit_status, type_of_unit, station_name, division, section, categorisation in query.order_by(Earning.date_of_receipt.desc().nullslast(), Earning.earning_key.desc()).all()]
    finally:
        session.close()


def list_commercial_contracts(q: str | None = None, station_code: str | None = None, policy: str | None = None, sub_category: str | None = None, allocation_code: str | None = None) -> list[dict[str, Any]]:
    session = SessionLocal()
    try:
        query = (
            session.query(CommercialContract, CommercialContractStationLink, Station.station_name, Station.division, Station.section, Station.categorisation)
            .join(CommercialContractStationLink, CommercialContractStationLink.contract_key == CommercialContract.contract_key, isouter=True)
            .join(Station, Station.station_code == CommercialContractStationLink.station_code, isouter=True)
        )
        if q:
            like = f"%{q}%"
            query = query.filter(
                (CommercialContract.contract_name.ilike(like))
                | (CommercialContract.licensee_name.ilike(like))
                | (CommercialContract.policy.ilike(like))
                | (CommercialContract.sub_category.ilike(like))
                | (CommercialContract.asset_scope.ilike(like))
                | (CommercialContract.allocation_code.ilike(like))
                | (CommercialContract.raw_station_value.ilike(like))
                | (CommercialContractStationLink.station_code.ilike(like))
            )
        if station_code and station_code != "All":
            query = query.filter(CommercialContractStationLink.station_code == station_code)
        if policy and policy != "All":
            query = query.filter(CommercialContract.policy == policy)
        if sub_category and sub_category != "All":
            query = query.filter(CommercialContract.sub_category == sub_category)
        if allocation_code and allocation_code != "All":
            query = query.filter(CommercialContract.allocation_code == allocation_code)

        rows = []
        seen_stationless = set()
        for contract, link, station_name, division, section, categorisation in query.order_by(CommercialContract.policy, CommercialContract.contract_name).all():
            if not link and contract.contract_key in seen_stationless:
                continue
            seen_stationless.add(contract.contract_key)
            rows.append({
                **row_to_dict(contract),
                "station_code": link.station_code if link else None,
                "station_name": station_name,
                "division": division,
                "section": section,
                "categorisation": categorisation,
                "station_match_status": "Station linked" if link and link.station_code else contract.station_match_status,
                "link_raw_station_value": link.raw_station_value if link else contract.raw_station_value,
                "match_type": link.match_type if link else None,
                "match_status": link.match_status if link else contract.station_match_status,
            })
        return rows
    finally:
        session.close()


def get_commercial_contract_detail(contract_key: int) -> dict[str, Any] | None:
    session = SessionLocal()
    try:
        contract = session.get(CommercialContract, contract_key)
        if not contract:
            return None
        links = (
            session.query(CommercialContractStationLink, Station.station_name, Station.division, Station.section, Station.categorisation)
            .join(Station, Station.station_code == CommercialContractStationLink.station_code, isouter=True)
            .filter(CommercialContractStationLink.contract_key == contract_key)
            .order_by(CommercialContractStationLink.station_code)
            .all()
        )
        payments = session.query(CommercialContractPayment).filter(CommercialContractPayment.contract_key == contract_key).order_by(CommercialContractPayment.payment_month).all()
        return {
            "contract": row_to_dict(contract),
            "station_links": [
                {
                    **row_to_dict(link),
                    "station_name": station_name,
                    "division": division,
                    "section": section,
                    "categorisation": categorisation,
                }
                for link, station_name, division, section, categorisation in links
            ],
            "payments": [row_to_dict(row) for row in payments],
            "payment_total": sum(to_money(row.amount_paid) for row in payments),
        }
    finally:
        session.close()


def get_contract_alerts(today: date | None = None, station_code: str | None = None) -> dict[str, Any]:
    """Combine catering and non-catering risk into one action-centre payload."""
    today = today or date.today()
    catering = list_units(station_code=station_code)
    commercial = list_commercial_contracts(station_code=station_code)
    commercial_report = get_commercial_contract_reports(today)
    commercial_risk = {row["contract_key"]: row for row in commercial_report.get("contract_risk", [])}
    rows: list[dict[str, Any]] = []

    for row in catering:
        if not row.get("notification_required"):
            continue
        rows.append({
            **row,
            "source_module": "catering",
            "contract_key": row.get("unit_no"),
            "contract_name": row.get("licensee_name") or row.get("unit_no"),
            "contract_type": row.get("type_of_unit") or "Catering",
            "station_code": row.get("station_code"),
            "license_fee": row.get("license_fee"),
        })

    for row in commercial:
        risk = commercial_risk.get(row.get("contract_key"), contract_risk(
            row.get("contract_upto"),
            station_match_status=row.get("station_match_status"),
            missing_fee=not to_money(row.get("annual_license_fee")),
            today=today,
        ))
        if not risk.get("notification_required"):
            continue
        rows.append({
            **row,
            **risk,
            "source_module": "commercial",
            "contract_type": row.get("sub_category") or row.get("policy") or "Commercial",
        })

    rows.sort(key=lambda item: (-int(item.get("risk_score") or 0), item.get("days_to_expiry") is None, item.get("contract_name") or ""))
    buckets = {
        "expired": 0,
        "due_today": 0,
        "due_5_days": 0,
        "due_10_days": 0,
        "due_30_days": 0,
        "due_50_days": 0,
        "due_90_days": 0,
        "payment_attention": 0,
    }
    for row in rows:
        bucket = row.get("expiry_alert_bucket")
        if bucket in buckets:
            buckets[bucket] += 1
        if row.get("pending_amount"):
            buckets["payment_attention"] += 1
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_alerts": len(rows),
        "critical": sum(1 for row in rows if row.get("risk_level") == "critical"),
        "high": sum(1 for row in rows if row.get("risk_level") == "high"),
        "buckets": buckets,
        "rows": rows,
    }


def get_action_centre(station_code: str | None = None, limit: int = 200, today: date | None = None) -> dict[str, Any]:
    """Return one normalized exception feed for web, Station 360, and mobile clients."""
    today = today or date.today()
    items: list[dict[str, Any]] = []

    contract_alerts = get_contract_alerts(today=today, station_code=station_code)
    for row in contract_alerts.get("rows", []):
        days = row.get("days_to_expiry")
        severity = "critical" if row.get("risk_level") == "critical" else "high" if row.get("risk_level") == "high" else "medium"
        items.append({
            "notification_key": f"contract:{row.get('source_module')}:{row.get('contract_key')}",
            "type": "contract_expiry" if days is not None else "payment_attention",
            "severity": severity,
            "station_code": row.get("station_code"),
            "record_id": row.get("contract_key"),
            "title": row.get("contract_name") or row.get("contract_key"),
            "message": "Expired" if days is not None and days < 0 else f"{days} days to expiry" if days is not None else "Payment or validity attention required",
            "days_to_expiry": days,
            "pending_amount": row.get("pending_amount", 0),
            "source_module": row.get("source_module"),
            "updated_at": row.get("updated_at"),
        })

    work_report = get_work_monitoring(station_code=station_code, today=today)
    for row in work_report.get("items", []):
        if not row.get("alert_type"):
            continue
        items.append({
            "notification_key": f"work:{row.get('project_id')}:{row.get('alert_type')}",
            "type": "work_tdc" if row.get("alert_type") != "contradiction" else "work_contradiction",
            "severity": "critical" if row.get("alert_type") == "tdc_overdue" else "high",
            "station_code": row.get("station_code"),
            "record_id": row.get("project_id"),
            "title": row.get("short_name_of_work") or row.get("project_id"),
            "message": row.get("alert_label") or "Work data requires review",
            "days_to_tdc": row.get("days_to_tdc"),
            "progress_percent": row.get("progress_percent"),
            "source_module": "works",
            "updated_at": row.get("updated_at"),
        })

    report = get_reports(today)
    for finding in report.get("inspections", {}).get("overdue_findings", []):
        if station_code and finding.get("station_code") != station_code.upper():
            continue
        items.append({
            "notification_key": f"finding:{finding.get('finding_id')}",
            "type": "inspection_action_overdue",
            "severity": "critical" if str(finding.get("severity", "")).lower() == "critical" else "high",
            "station_code": finding.get("station_code"),
            "record_id": finding.get("finding_id"),
            "title": finding.get("title") or "Inspection finding",
            "message": f"{finding.get('days_overdue', 0)} days overdue",
            "days_overdue": finding.get("days_overdue"),
            "source_module": "inspections",
            "updated_at": finding.get("updated_at"),
        })

    for row in list_passenger_amenities(kind="summary", station_code=station_code):
        missing = []
        if not row.get("platform_detail_count"):
            missing.append("platform details")
        if row.get("wheel_chairs") is None:
            missing.append("wheelchair record")
        if not row.get("fob_details"):
            missing.append("FOB/access details")
        if not missing:
            continue
        items.append({
            "notification_key": f"amenity:{row.get('station_code')}:{','.join(missing)}",
            "type": "amenity_incomplete",
            "severity": "medium",
            "station_code": row.get("station_code"),
            "record_id": row.get("station_code"),
            "title": row.get("station_name") or row.get("station_code"),
            "message": "Missing " + ", ".join(missing),
            "source_module": "amenities",
            "updated_at": None,
        })

    data_centre = get_data_centre_status()
    for failure in data_centre.get("failures", []):
        items.append({
            "notification_key": f"sync:{failure.get('resource')}:{failure.get('at')}",
            "type": "sync_failure",
            "severity": "critical",
            "station_code": None,
            "record_id": failure.get("resource"),
            "title": f"{failure.get('resource', 'Data')} synchronization failed",
            "message": failure.get("message") or "Source refresh failed",
            "source_module": "data-centre",
            "updated_at": failure.get("at"),
        })
    for change in data_centre.get("recent_sync", [])[:10]:
        if not re.search(r"import|replace|sync", clean(change.get("action")), flags=re.I):
            continue
        items.append({
            "notification_key": f"change:{change.get('resource')}:{change.get('at')}",
            "type": "source_changed",
            "severity": "info",
            "station_code": None,
            "record_id": change.get("resource"),
            "title": f"{change.get('resource', 'Source')} refreshed",
            "message": change.get("details") or "New source changes available",
            "source_module": "data-centre",
            "updated_at": change.get("at"),
        })

    unique: dict[str, dict[str, Any]] = {}
    for item in items:
        unique[item["notification_key"]] = item
    severity_order = {"critical": 0, "high": 1, "medium": 2, "info": 3}
    ordered = sorted(unique.values(), key=lambda item: (severity_order.get(item.get("severity"), 9), str(item.get("title") or "")))[:max(1, min(limit, 1000))]
    return {
        "as_of": today.isoformat(),
        "total": len(ordered),
        "critical": sum(1 for item in ordered if item.get("severity") == "critical"),
        "high": sum(1 for item in ordered if item.get("severity") == "high"),
        "medium": sum(1 for item in ordered if item.get("severity") == "medium"),
        "items": ordered,
    }


def get_commercial_contract_reports(today: date | None = None) -> dict[str, Any]:
    today = today or date.today()
    next_30 = today + timedelta(days=30)
    next_90 = today + timedelta(days=90)
    session = SessionLocal()
    try:
        contracts = session.query(CommercialContract).all()
        payments = session.query(CommercialContractPayment, CommercialContract).join(CommercialContract, CommercialContract.contract_key == CommercialContractPayment.contract_key).all()
        links = session.query(CommercialContractStationLink).all()

        by_policy: dict[str, int] = {}
        by_policy_value: dict[str, int] = {}
        by_sub_category: dict[str, int] = {}
        by_asset_scope: dict[str, int] = {}
        expiry_alerts = []
        expiry_buckets = {
            "expired": 0,
            "today": 0,
            "next_5_days": 0,
            "next_10_days": 0,
            "next_30_days": 0,
            "next_50_days": 0,
            "next_90_days": 0,
        }
        unmatched = []
        pending_by_contract: dict[int, int] = {}
        for payment, _ in payments:
            pending_by_contract[payment.contract_key] = pending_by_contract.get(payment.contract_key, 0) + max(
                0,
                to_money(payment.amount_due) - to_money(payment.amount_paid),
            )
        contract_risk_rows = []
        for contract in contracts:
            policy_key = clean(contract.policy) or "Unclassified"
            sub_key = clean(contract.sub_category) or "Unclassified"
            scope_key = clean(contract.asset_scope) or "Unclassified"
            annual = to_money(contract.annual_license_fee)
            by_policy[policy_key] = by_policy.get(policy_key, 0) + 1
            by_policy_value[policy_key] = by_policy_value.get(policy_key, 0) + annual
            by_sub_category[sub_key] = by_sub_category.get(sub_key, 0) + 1
            by_asset_scope[scope_key] = by_asset_scope.get(scope_key, 0) + 1
            pending_amount = pending_by_contract.get(contract.contract_key, 0)
            risk = contract_risk(
                contract.contract_upto,
                pending_amount=pending_amount,
                station_match_status=contract.station_match_status,
                missing_fee=not annual,
                today=today,
            )
            contract_risk_rows.append({
                "contract_key": contract.contract_key,
                "contract_name": contract.contract_name,
                "policy": contract.policy,
                "sub_category": contract.sub_category,
                "contract_upto": contract.contract_upto,
                **risk,
            })
            days_to_expiry = risk["days_to_expiry"]
            if days_to_expiry is not None and days_to_expiry <= 90:
                if days_to_expiry < 0:
                    bucket = "expired"
                elif days_to_expiry == 0:
                    bucket = "today"
                elif days_to_expiry <= 5:
                    bucket = "next_5_days"
                elif days_to_expiry <= 10:
                    bucket = "next_10_days"
                elif days_to_expiry <= 30:
                    bucket = "next_30_days"
                elif days_to_expiry <= 50:
                    bucket = "next_50_days"
                else:
                    bucket = "next_90_days"
                expiry_buckets[bucket] += 1
                expiry_alerts.append({
                    "contract_key": contract.contract_key,
                    "contract_name": contract.contract_name,
                    "licensee_name": contract.licensee_name,
                    "policy": contract.policy,
                    "sub_category": contract.sub_category,
                    "contract_upto": contract.contract_upto,
                    "days_to_expiry": days_to_expiry,
                    "alert_bucket": bucket,
                    "risk_score": risk["risk_score"],
                    "risk_level": risk["risk_level"],
                    "expiry_alert_bucket": risk["expiry_alert_bucket"],
                })
            if contract.station_match_status in {"unmatched", "asset_scope"}:
                unmatched.append({
                    "contract_key": contract.contract_key,
                    "contract_name": contract.contract_name,
                    "raw_station_value": contract.raw_station_value,
                    "policy": contract.policy,
                    "sub_category": contract.sub_category,
                    "asset_scope": contract.asset_scope,
                    "station_match_status": contract.station_match_status,
                })

        by_month: dict[str, int] = {}
        pending_like = []
        for payment, contract in payments:
            by_month[payment.payment_month] = by_month.get(payment.payment_month, 0) + to_money(payment.amount_paid)
            if to_money(payment.amount_due) > to_money(payment.amount_paid):
                pending_like.append({
                    **row_to_dict(payment),
                    "contract_name": contract.contract_name,
                    "licensee_name": contract.licensee_name,
                    "policy": contract.policy,
                    "sub_category": contract.sub_category,
                    "pending_amount": to_money(payment.amount_due) - to_money(payment.amount_paid),
                })

        station_link_counts: dict[str, int] = {}
        for link in links:
            key = clean(link.station_code) or clean(link.match_status) or "Asset Scope"
            station_link_counts[key] = station_link_counts.get(key, 0) + 1

        return {
            "total_contracts": len(contracts),
            "linked_station_rows": sum(1 for link in links if link.station_code),
            "asset_scope_rows": sum(1 for link in links if not link.station_code),
            "annual_license_fee": sum(to_money(row.annual_license_fee) for row in contracts),
            "recorded_payments": sum(to_money(payment.amount_paid) for payment, _ in payments),
            "by_policy": [{"label": key, "value": value, "amount": by_policy_value.get(key, 0)} for key, value in sorted(by_policy.items(), key=lambda item: item[1], reverse=True)],
            "by_sub_category": [{"label": key, "value": value} for key, value in sorted(by_sub_category.items(), key=lambda item: item[1], reverse=True)[:20]],
            "by_asset_scope": [{"label": key, "value": value} for key, value in sorted(by_asset_scope.items(), key=lambda item: item[1], reverse=True)],
            "by_month": [{"label": key, "value": value} for key, value in sorted(by_month.items())[-24:]],
            "by_station": [{"label": key, "value": value} for key, value in sorted(station_link_counts.items(), key=lambda item: item[1], reverse=True)[:20]],
            "expiry_alerts": sorted(expiry_alerts, key=lambda row: row["days_to_expiry"]),
            "expiry_buckets": expiry_buckets,
            "pending_payments": pending_like,
            "contract_risk": sorted(contract_risk_rows, key=lambda row: (-row["risk_score"], row["contract_name"])),
            "unmatched_or_asset_scope": unmatched,
        }
    finally:
        session.close()


def _fob_access_details(
    infra: StationInfra | None,
    status: StationPlatformExtensionStatus | None,
) -> list[str]:
    """Convert PA Infra Lifts & Ramp flags into station-facing access modes."""
    if not infra and not status:
        return []
    modes: list[str] = []
    fob_text = normalize(infra.fob_details if infra else "")
    if status and status.fob_without:
        modes.append("FOB not recorded")
    elif fob_text and fob_text not in {"-", "none", "no"}:
        modes.append("FOB")
        if status and status.fob_ramp_available:
            modes.append("Ramp available")
        if status and status.lift_available:
            modes.append("Lift available")
        if not status or not status.fob_ramp_available and not status.lift_available:
            modes.append("Stairs")
    if status and status.fob_wip:
        modes.append("FOB work in progress")
    if status and status.ramp_proposed:
        modes.append("Ramp proposed")
    if status and status.lift_proposed:
        modes.append("Lift proposed")
    if status and status.not_feasible_lift_ramp:
        modes.append("Lift / ramp not feasible")
    return list(dict.fromkeys(modes))


def get_station_detail(station_code: str) -> dict[str, Any] | None:
    code = clean(station_code).upper()
    session = SessionLocal()
    try:
        station = session.get(Station, code)
        if not station:
            return None

        station_row = row_to_dict(station)
        units = list_units(station_code=code)
        earnings = list_earnings(station_code=code)
        earnings_by_unit: dict[str, list[dict[str, Any]]] = {}
        for earning in earnings:
            if earning.get("unit_no"):
                earnings_by_unit.setdefault(earning["unit_no"], []).append(earning)
        today = date.today()

        def with_validity(row: dict[str, Any], start_keys: tuple[str, ...], end_keys: tuple[str, ...]) -> dict[str, Any]:
            enriched = dict(row)
            start_value = next((row.get(key) for key in start_keys if row.get(key)), None)
            end_value = next((row.get(key) for key in end_keys if row.get(key)), None)
            start_date = parse_date_value(start_value)
            end_date = parse_date_value(end_value)
            enriched["valid_from"] = start_date.isoformat() if start_date else None
            enriched["valid_to"] = end_date.isoformat() if end_date else None
            days = (end_date - today).days if end_date else None
            enriched["days_to_expiry"] = days
            enriched["renewal_state"] = (
                "Date unavailable" if days is None else
                "Expired" if days < 0 else
                "Due within 7 days" if days <= 7 else
                "Renewal due within 30 days" if days <= 30 else
                "Renewal upcoming" if days <= 90 else
                "Active"
            )
            return enriched

        contracts = []
        for unit in units:
            available = is_available_unit(unit)
            unit_earnings = [] if available else earnings_by_unit.get(unit.get("unit_no"), [])
            enriched = {
                **unit,
                "unit_status": "Available" if available else unit.get("unit_status"),
                "availability_remarks": unit.get("remarks") or (unit.get("unit_status") if available else None),
                "licensee_name": None if available else unit.get("licensee_name") or next(
                    (row.get("licensee_name") for row in unit_earnings if clean(row.get("licensee_name"))),
                    None,
                ),
                "earnings": unit_earnings,
                "earnings_total": sum(to_money(row.get("amount")) for row in unit_earnings),
                "payment_total": sum(to_money(row.get("amount")) for row in unit_earnings),
                "pending_receipts": sum(1 for row in unit_earnings if "pending" in normalize(row.get("receipt_type"))),
            }
            contracts.append(with_validity(enriched, ("contract_from",), ("contract_to",)))

        works = list_works(station_code=code)
        commercial_contracts = list_commercial_contracts(station_code=code)
        publicity_contracts = [
            row for row in list_registry_contracts()
            if any(asset.get("station_code") == code for asset in row.get("assets", []))
        ]
        monthly_metrics = (
            session.query(StationMonthlyMetric)
            .filter(StationMonthlyMetric.station_code == code)
            .order_by(StationMonthlyMetric.metric_month.desc())
            .all()
        )
        commercial_keys = {
            row.get("contract_key")
            for row in commercial_contracts
            if row.get("contract_key") is not None
        }
        commercial_payments: dict[Any, list[dict[str, Any]]] = {}
        if commercial_keys:
            payment_rows = (
                session.query(CommercialContractPayment)
                .filter(CommercialContractPayment.contract_key.in_(commercial_keys))
                .order_by(CommercialContractPayment.payment_month)
                .all()
            )
            for payment in payment_rows:
                commercial_payments.setdefault(payment.contract_key, []).append(row_to_dict(payment))
        for contract in commercial_contracts:
            contract.update(with_validity(contract, ("contract_period_from",), ("contract_upto",)))
            payments = commercial_payments.get(contract.get("contract_key"), [])
            contract["payments"] = payments
            contract["payment_total"] = sum(to_money(row.get("amount_paid")) for row in payments)
        infra = session.query(StationInfra).filter(StationInfra.station_code == code).one_or_none()
        platforms = session.query(PlatformDetail).filter(PlatformDetail.station_code == code).order_by(PlatformDetail.platform).all()
        wheelchairs = session.query(WheelChairAvailability).filter(WheelChairAvailability.station_code == code).one_or_none()
        trolley = session.query(TrolleyPath).filter(TrolleyPath.station_code == code).one_or_none()
        pa_works = session.query(PassengerAmenityWork).filter(PassengerAmenityWork.station_code == code).order_by(PassengerAmenityWork.work_type, PassengerAmenityWork.work_name).all()
        pf_extension = session.query(StationPlatformExtensionStatus).filter(StationPlatformExtensionStatus.station_code == code).one_or_none()
        norms = session.query(AmenityNorm).filter(AmenityNorm.category == station.categorisation).order_by(AmenityNorm.amenity, AmenityNorm.norm).all() if station.categorisation else []
        inspections = (
            session.query(Inspection)
            .filter(Inspection.station_code == code)
            .order_by(Inspection.updated_at.desc())
            .limit(20)
            .all()
        )
        inspection_findings = (
            session.query(InspectionFinding)
            .filter(
                InspectionFinding.station_code == code,
                InspectionFinding.status.notin_(["verified", "closed"]),
            )
            .order_by(InspectionFinding.updated_at.desc())
            .limit(100)
            .all()
        )

        platform_lengths = [row.length_m for row in platforms if row.length_m is not None]
        amenities = {
            "infra": row_to_dict(infra) if infra else None,
            "platforms": [row_to_dict(row) for row in platforms],
            "wheelchairs": row_to_dict(wheelchairs) if wheelchairs else None,
            "trolley": row_to_dict(trolley) if trolley else None,
            "pa_works": [row_to_dict(row) for row in pa_works],
            "pf_extension_status": row_to_dict(pf_extension) if pf_extension else None,
            "fob_access": _fob_access_details(infra, pf_extension),
            "norms": [row_to_dict(row) for row in norms],
        }
        amenity_summary = {
            "platforms": len(platforms),
            "platform_count_declared": infra.platform_count if infra else station.number_of_platforms,
            "total_platform_length": sum(platform_lengths),
            "shortest_platform": min(platform_lengths) if platform_lengths else None,
            "longest_platform": max(platform_lengths) if platform_lengths else None,
            "wheel_chairs": wheelchairs.available_good_condition if wheelchairs else None,
            "trolley_path": trolley.trolley_path if trolley else None,
            "fob_details": infra.fob_details if infra else None,
            "pa_works": len(pa_works),
            "open_pa_works": sum(1 for row in pa_works if "complete" not in normalize(row.progress)),
            "norms": len(norms),
            "pf_extension_wip": bool(pf_extension and pf_extension.pf_extension_wip),
            "pf_extension_proposed": bool(pf_extension and pf_extension.pf_extension_proposed),
            "raising_extension_proposed": bool(pf_extension and pf_extension.raising_extension_proposed),
            "platform_extension_work_proposed": bool(pf_extension and pf_extension.platform_extension_work_proposed),
            "ramp_feasible": bool(pf_extension and pf_extension.ramp_feasible),
            "lift_proposed": bool(pf_extension and pf_extension.lift_proposed),
            "ramp_proposed": bool(pf_extension and pf_extension.ramp_proposed),
            "not_feasible_lift_ramp": bool(pf_extension and pf_extension.not_feasible_lift_ramp),
        }
        recorded_amenities = dict(amenities)
        recorded_amenities["norms"] = []
        recorded_amenity_text = normalize(str(recorded_amenities))
        compliance_groups: dict[str, dict[str, int]] = {}
        missing_norms = []
        matched_norms = 0
        for norm_row in norms:
            category = clean(norm_row.category) or "Unclassified"
            group = compliance_groups.setdefault(category, {"total": 0, "matched": 0})
            group["total"] += 1
            labels = [clean(norm_row.amenity), clean(norm_row.norm)]
            labels = [normalize(label) for label in labels if len(normalize(label)) >= 3]
            matched = any(label in recorded_amenity_text for label in labels)
            if matched:
                group["matched"] += 1
                matched_norms += 1
            else:
                missing_norms.append({
                    "norm_key": norm_row.norm_key,
                    "category": category,
                    "amenity": norm_row.amenity,
                    "norm": norm_row.norm,
                    "quantity": norm_row.norm_quantity,
                })
        amenity_compliance = {
            "total_norms": len(norms),
            "matched_norms": matched_norms,
            "compliance_percent": round((matched_norms / len(norms)) * 100, 1) if norms else None,
            "by_category": [
                {
                    "category": category,
                    "total": values["total"],
                    "matched": values["matched"],
                    "compliance_percent": round((values["matched"] / values["total"]) * 100, 1) if values["total"] else 0,
                }
                for category, values in sorted(compliance_groups.items())
            ],
            "missing": missing_norms[:200],
        }
        work_ids = [row.get("project_id") for row in works if row.get("project_id")]
        progress_by_work: dict[str, list[dict[str, Any]]] = {}
        if work_ids:
            progress_rows = (
                session.query(WorkProgressUpdate)
                .filter(WorkProgressUpdate.project_id.in_(work_ids))
                .order_by(WorkProgressUpdate.update_date.desc(), WorkProgressUpdate.progress_id.desc())
                .all()
            )
            for progress in progress_rows:
                progress_by_work.setdefault(progress.project_id, []).append(row_to_dict(progress))
        for work in works:
            work["progress_updates"] = progress_by_work.get(work.get("project_id"), [])
        open_works = [
            row for row in works
            if not re.search(r"complete|done", normalize(row.get("status")))
        ]
        amenity_flags = []
        if not infra:
            amenity_flags.append({"key": "infra", "label": "Station infra details missing", "severity": "high"})
        if not platforms:
            amenity_flags.append({"key": "platforms", "label": "Platform-wise details missing", "severity": "medium"})
        if not wheelchairs:
            amenity_flags.append({"key": "wheelchair", "label": "Wheelchair availability not recorded", "severity": "medium"})
        if not norms:
            amenity_flags.append({"key": "norms", "label": "Category norms not linked", "severity": "medium"})
        if summary_open := amenity_summary["open_pa_works"]:
            amenity_flags.append({"key": "pa_works", "label": f"{summary_open} passenger amenity works open", "severity": "high"})

        contract_alerts = []
        for contract in [*contracts, *commercial_contracts, *publicity_contracts]:
            days = contract.get("days_to_expiry")
            pending = contract.get("pending_receipts", 0) or 0
            if (days is not None and days <= 90) or pending:
                contract_alerts.append({
                    "contract_key": contract.get("contract_key") or contract.get("unit_no"),
                    "contract_name": contract.get("contract_name") or contract.get("licensee_name") or contract.get("unit_no"),
                    "days_to_expiry": days,
                    "pending_receipts": pending,
                    "severity": "critical" if days is not None and days <= 10 else "warning",
                })

        timeline = []
        for row in works:
            timeline.append({
                "type": "work",
                "date": row.get("updated_at") or row.get("date_of_sanction"),
                "title": row.get("short_name") or row.get("work_name") or row.get("project_id"),
                "status": row.get("status"),
                "record_id": row.get("project_id"),
            })
        for row in earnings:
            timeline.append({
                "type": "payment",
                "date": row.get("date_of_receipt") or row.get("mr_date"),
                "title": row.get("payment_head") or row.get("receipt_key"),
                "status": row.get("receipt_type"),
                "record_id": row.get("receipt_key") or row.get("earning_key"),
            })
        for row in inspections:
            timeline.append({
                "type": "inspection",
                "date": row.updated_at,
                "title": f"Inspection by {row.inspector_name}",
                "status": row.status,
                "record_id": row.inspection_id,
            })
        for row in inspection_findings:
            timeline.append({
                "type": "deficiency",
                "date": row.updated_at,
                "title": row.title,
                "status": row.status,
                "record_id": row.finding_id,
            })
        timeline.sort(key=lambda row: str(row.get("date") or ""), reverse=True)
        inspection_comparison: dict[str, Any] = {
            "available": len(inspections) >= 2,
            "current": None,
            "previous": None,
            "score_delta": None,
            "changed_items": [],
        }
        if len(inspections) >= 2:
            current_inspection, previous_inspection = inspections[0], inspections[1]
            current_responses = {
                (row.section_code, row.question_code, row.platform): row
                for row in session.query(InspectionResponse).filter(
                    InspectionResponse.inspection_id == current_inspection.inspection_id,
                ).all()
            }
            previous_responses = {
                (row.section_code, row.question_code, row.platform): row
                for row in session.query(InspectionResponse).filter(
                    InspectionResponse.inspection_id == previous_inspection.inspection_id,
                ).all()
            }
            changed_items = []
            for key in sorted(
                set(current_responses) | set(previous_responses),
                key=lambda item: tuple(str(part or "") for part in item),
            ):
                current = current_responses.get(key)
                previous = previous_responses.get(key)
                current_value = current.response_value if current else None
                previous_value = previous.response_value if previous else None
                if current_value == previous_value:
                    continue
                changed_items.append({
                    "section_code": key[0],
                    "question_code": key[1],
                    "platform": key[2],
                    "previous": previous_value,
                    "current": current_value,
                    "severity": (current.severity if current else previous.severity) if (current or previous) else None,
                    "previous_remarks": previous.remarks if previous else None,
                    "current_remarks": current.remarks if current else None,
                })
            inspection_comparison = {
                "available": True,
                "current": {
                    "inspection_id": current_inspection.inspection_id,
                    "status": current_inspection.status,
                    "score": current_inspection.score,
                    "updated_at": current_inspection.updated_at.isoformat() if current_inspection.updated_at else None,
                },
                "previous": {
                    "inspection_id": previous_inspection.inspection_id,
                    "status": previous_inspection.status,
                    "score": previous_inspection.score,
                    "updated_at": previous_inspection.updated_at.isoformat() if previous_inspection.updated_at else None,
                },
                "score_delta": (
                    current_inspection.score - previous_inspection.score
                    if current_inspection.score is not None and previous_inspection.score is not None
                    else None
                ),
                "changed_items": changed_items[:300],
            }
        return {
            "station": station_row,
            "contracts": contracts,
            "units": units,
            "earnings": earnings,
            "works": works,
            "commercial_contracts": commercial_contracts,
            "publicity_contracts": publicity_contracts,
            "monthly_metrics": [row_to_dict(row) for row in monthly_metrics],
            "latest_monthly_metric": row_to_dict(monthly_metrics[0]) if monthly_metrics else None,
            "amenities": amenities,
            "amenity_summary": amenity_summary,
            "amenity_compliance": amenity_compliance,
            "action_centre": {
                "contract_alerts": contract_alerts,
                "open_works": open_works,
                "amenity_flags": amenity_flags,
                "open_findings": [row_to_dict(row) for row in inspection_findings],
                "inspections": [row_to_dict(row) for row in inspections],
                "inspection_comparison": inspection_comparison,
                "timeline": timeline[:100],
            },
        }
    finally:
        session.close()


def get_reports(today: date | None = None) -> dict[str, Any]:
    today = today or date.today()
    current_month_end = month_end(today)
    next_7 = today + timedelta(days=7)
    next_30 = today + timedelta(days=30)
    next_90 = today + timedelta(days=90)
    three_month_start = date(today.year, today.month, 1) - timedelta(days=62)

    session = SessionLocal()
    try:
        commercial_station_filter = (
            Station.is_active.is_(True),
            func.lower(func.trim(Station.categorisation)).notin_(('', 'test', 'non-commercial')),
        )
        stations = session.query(Station).filter(*commercial_station_filter).all()
        units = session.query(Unit, Station.station_name, Station.division, Station.section).join(Station, Station.station_code == Unit.station_code, isouter=True).all()
        earnings = session.query(Earning).filter(
            (Earning.earning_scope.is_(None)) | (Earning.earning_scope != "tender_emd")
        ).all()
        works = session.query(
            Work,
            WorkLink,
            Station.station_name,
            Station.sr_den,
            Station.cmi,
        ).join(WorkLink, WorkLink.project_id == Work.project_id, isouter=True).join(Station, Station.station_code == WorkLink.station_code, isouter=True).all()
        inspections = session.query(Inspection).all()
        findings = session.query(InspectionFinding).all()

        by_unit: dict[str, list[Earning]] = {}
        for earning in earnings:
            if earning.unit_no:
                by_unit.setdefault(earning.unit_no, []).append(earning)

        station_codes = {station.station_code for station in stations}
        station_names = {station.station_code: station.station_name for station in stations}
        unit_codes = {unit.unit_no for unit, *_ in units}
        station_with_units = {unit.station_code for unit, *_ in units if unit.station_code}
        station_with_works = {link.station_code for _, link, *_ in works if link and link.station_code}
        station_with_earnings = {earning.station_code for earning in earnings if earning.station_code}

        by_station_category: dict[str, int] = {}
        by_division: dict[str, int] = {}
        by_section: dict[str, int] = {}
        for station in stations:
            by_station_category[clean(station.categorisation) or "Unknown"] = by_station_category.get(clean(station.categorisation) or "Unknown", 0) + 1
            by_division[clean(station.division) or "Unknown"] = by_division.get(clean(station.division) or "Unknown", 0) + 1
            by_section[clean(station.section) or "Unknown"] = by_section.get(clean(station.section) or "Unknown", 0) + 1

        active_units = [(unit, station_name, division, section) for unit, station_name, division, section in units if is_active_status(unit.unit_status)]
        inactive_units = len(units) - len(active_units)
        by_status: dict[str, int] = {}
        by_type: dict[str, int] = {}
        by_unit_category: dict[str, int] = {}
        for unit, *_ in units:
            by_status[clean(unit.unit_status) or "Unknown"] = by_status.get(clean(unit.unit_status) or "Unknown", 0) + 1
            by_type[clean(unit.type_of_unit) or "Unknown"] = by_type.get(clean(unit.type_of_unit) or "Unknown", 0) + 1
            by_unit_category[clean(unit.station_category) or "Unknown"] = by_unit_category.get(clean(unit.station_category) or "Unknown", 0) + 1

        license_earnings = [row for row in earnings if is_license_fee_row(row)]
        pending_receipts = [row for row in earnings if "pending" in normalize(row.receipt_type)]
        current_month_license = [
            row for row in license_earnings
            if (parsed := parse_date_value(row.date_of_receipt or row.mr_date)) and parsed.year == today.year and parsed.month == today.month
        ]
        recent_license = [
            row for row in license_earnings
            if (parsed := parse_date_value(row.date_of_receipt or row.mr_date)) and parsed >= three_month_start
        ]

        earnings_by_month: dict[str, int] = {}
        earnings_by_head: dict[str, int] = {}
        for row in earnings:
            receipt_date = parse_date_value(row.date_of_receipt or row.mr_date)
            month_key = receipt_date.strftime("%Y-%m") if receipt_date else "Unknown"
            head_key = clean(row.payment_head) or "Unknown"
            earnings_by_month[month_key] = earnings_by_month.get(month_key, 0) + to_money(row.amount)
            earnings_by_head[head_key] = earnings_by_head.get(head_key, 0) + to_money(row.amount)

        works_by_status: dict[str, int] = {}
        works_by_scope: dict[str, int] = {}
        works_by_section: dict[str, int] = {}
        works_by_station: dict[str, int] = {}
        works_by_sr_den: dict[str, int] = {}
        works_by_cmi: dict[str, int] = {}
        works_by_allocation: dict[str, int] = {}
        works_by_year: dict[str, int] = {}
        works_seen: set[str] = set()
        completed_work_count = 0
        pending_work_count = 0
        delay_alerts: list[dict[str, Any]] = []
        work_contradictions: list[dict[str, Any]] = []

        def progress_percent(value: Any) -> int | None:
            match = re.search(r"(\d+(?:\.\d+)?)\s*%?", clean(value))
            if not match:
                return None
            try:
                return round(float(match.group(1)))
            except ValueError:
                return None

        def canonical_label(values: dict[str, int], value: str) -> str:
            """Merge case-only variants without changing the first source label."""
            for existing in values:
                if existing.casefold() == value.casefold():
                    return existing
            return value

        for work, link, station_name, sr_den, cmi in works:
            if work.project_id not in works_seen:
                works_seen.add(work.project_id)
                status_key = clean(work.status) or "Unknown"
                section_key = canonical_label(works_by_section, clean(work.section) or "Unknown")
                works_by_status[status_key] = works_by_status.get(status_key, 0) + 1
                works_by_section[section_key] = works_by_section.get(section_key, 0) + 1
                sr_den_key = clean(sr_den) or "Unassigned"
                cmi_key = clean(cmi) or "Unassigned"
                allocation_key = clean(work.allocation) or "Unassigned"
                year_key = clean(work.year_of_sanction) or "Unknown"
                works_by_sr_den[sr_den_key] = works_by_sr_den.get(sr_den_key, 0) + 1
                works_by_cmi[cmi_key] = works_by_cmi.get(cmi_key, 0) + 1
                works_by_allocation[allocation_key] = works_by_allocation.get(allocation_key, 0) + 1
                works_by_year[year_key] = works_by_year.get(year_key, 0) + 1
                if re.search(r"complete|done", status_key, flags=re.I):
                    completed_work_count += 1
                else:
                    pending_work_count += 1
                percentage = progress_percent(work.physical_progress)
                is_completed = bool(re.search(r"complete|done", status_key, flags=re.I))
                if is_completed and percentage is not None and percentage < 100:
                    work_contradictions.append({
                        "project_id": work.project_id,
                        "problem": "Completed status with progress below 100%",
                        "status": work.status,
                        "progress_percent": percentage,
                        "station_code": link.station_code if link else None,
                    })
                if not is_completed and percentage == 100:
                    work_contradictions.append({
                        "project_id": work.project_id,
                        "problem": "100% progress with an open status",
                        "status": work.status,
                        "progress_percent": percentage,
                        "station_code": link.station_code if link else None,
                    })
                tdc_date = parse_date_value(work.tdc)
                if tdc_date and tdc_date < today and not is_completed:
                    delay_alerts.append({
                        "project_id": work.project_id,
                        "short_name_of_work": work.short_name_of_work,
                        "status": work.status,
                        "progress_percent": percentage,
                        "tdc": work.tdc,
                        "days_overdue": (today - tdc_date).days,
                        "station_code": link.station_code if link else None,
                        "station_name": station_name,
                        "sr_den": sr_den,
                        "cmi": cmi,
                    })
            scope_key = clean(link.scope_type) if link else "Unlinked"
            station_key = clean(link.station_code) if link and link.station_code else clean(station_name) or "Unlinked"
            works_by_scope[scope_key or "Unknown"] = works_by_scope.get(scope_key or "Unknown", 0) + 1
            works_by_station[station_key] = works_by_station.get(station_key, 0) + 1

        alerts: list[dict[str, Any]] = []
        bucket_counts = {
            "overdue": 0,
            "due_this_month": 0,
            "due_next_7_days": 0,
            "due_next_30_days": 0,
            "due_next_90_days": 0,
            "contract_expired": 0,
            "contract_due_today": 0,
            "contract_due_5_days": 0,
            "contract_due_10_days": 0,
            "contract_due_30_days": 0,
            "contract_due_50_days": 0,
            "contract_due_90_days": 0,
            "needs_review": 0,
        }
        estimated_overdue_amount = 0

        for unit, station_name, division, section in active_units:
            unit_license_rows = [row for row in by_unit.get(unit.unit_no, []) if is_license_fee_row(row)]
            paid_through_values = [
                parsed for row in unit_license_rows
                if (parsed := parse_date_value(row.period_to or row.date_of_receipt or row.mr_date))
                and "pending" not in normalize(row.receipt_type)
            ]
            if sheet_paid_upto := parse_date_value(unit.paid_upto):
                paid_through_values.append(sheet_paid_upto)
            last_paid_through = max(paid_through_values) if paid_through_values else None
            contract_to = parse_date_value(unit.contract_to)
            license_fee_amount = to_money(unit.license_fee)

            if last_paid_through is None:
                months_pending = 1
                alert_bucket = "needs_review"
            elif last_paid_through < today:
                months_pending = month_delta(last_paid_through + timedelta(days=1), current_month_end)
                alert_bucket = "overdue"
            elif last_paid_through <= current_month_end:
                months_pending = 0
                alert_bucket = "due_this_month"
            else:
                months_pending = 0
                alert_bucket = None

            days_to_contract_end = (contract_to - today).days if contract_to else None
            if days_to_contract_end is not None and days_to_contract_end < 0:
                contract_bucket = "contract_expired"
            elif days_to_contract_end == 0:
                contract_bucket = "contract_due_today"
            elif days_to_contract_end is not None and days_to_contract_end <= 5:
                contract_bucket = "contract_due_5_days"
            elif days_to_contract_end is not None and days_to_contract_end <= 10:
                contract_bucket = "contract_due_10_days"
            elif days_to_contract_end is not None and days_to_contract_end <= 30:
                contract_bucket = "contract_due_30_days"
            elif days_to_contract_end is not None and days_to_contract_end <= 50:
                contract_bucket = "contract_due_50_days"
            elif days_to_contract_end is not None and days_to_contract_end <= 90:
                contract_bucket = "contract_due_90_days"
            else:
                contract_bucket = None

            if contract_bucket:
                bucket_counts[contract_bucket] += 1

            legacy_contract_bucket = None
            if days_to_contract_end is not None and days_to_contract_end < 0:
                legacy_contract_bucket = "overdue"
            elif days_to_contract_end is not None and 0 <= days_to_contract_end <= 7:
                legacy_contract_bucket = "due_next_7_days"
            elif days_to_contract_end is not None and 0 < days_to_contract_end <= 30:
                legacy_contract_bucket = "due_next_30_days"
            elif days_to_contract_end is not None and 30 < days_to_contract_end <= 90:
                legacy_contract_bucket = "due_next_90_days"

            final_bucket = alert_bucket or legacy_contract_bucket
            if not final_bucket:
                continue

            bucket_counts[final_bucket] = bucket_counts.get(final_bucket, 0) + 1
            estimated_pending = months_pending * license_fee_amount
            if final_bucket == "overdue":
                estimated_overdue_amount += estimated_pending

            alerts.append({
                "unit_no": unit.unit_no,
                "station_code": unit.station_code,
                "station_name": station_name,
                "division": division,
                "section": section,
                "licensee_name": unit.licensee_name,
                "type_of_unit": unit.type_of_unit,
                "unit_status": unit.unit_status,
                "license_fee": unit.license_fee,
                "license_fee_amount": license_fee_amount,
                "contract_to": unit.contract_to,
                "last_paid_through": last_paid_through.isoformat() if last_paid_through else None,
                "days_to_contract_end": days_to_contract_end,
                "contract_alert_bucket": contract_bucket,
                "months_pending": months_pending,
                "estimated_pending_amount": estimated_pending,
                "alert_bucket": final_bucket,
            })

        bucket_order = {
            "overdue": 0,
            "needs_review": 1,
            "due_this_month": 2,
            "due_next_7_days": 3,
            "due_next_30_days": 4,
            "due_next_90_days": 5,
        }
        alerts.sort(key=lambda row: (bucket_order.get(row["alert_bucket"], 9), row["days_to_contract_end"] is None, row["days_to_contract_end"] or 99999, row["unit_no"] or ""))
        commercial_reports = get_commercial_contract_reports(today)

        inspection_by_status: dict[str, int] = {}
        finding_by_status: dict[str, int] = {}
        finding_by_severity: dict[str, int] = {}
        finding_by_department: dict[str, int] = {}
        overdue_findings: list[dict[str, Any]] = []
        open_findings: list[dict[str, Any]] = []
        for inspection in inspections:
            status_key = clean(inspection.status) or "Unknown"
            inspection_by_status[status_key] = inspection_by_status.get(status_key, 0) + 1
        for finding in findings:
            status_key = clean(finding.status) or "Unknown"
            severity_key = clean(finding.severity) or "Unknown"
            department_key = clean(finding.responsible_party) or "Unassigned"
            finding_by_status[status_key] = finding_by_status.get(status_key, 0) + 1
            finding_by_severity[severity_key] = finding_by_severity.get(severity_key, 0) + 1
            finding_by_department[department_key] = finding_by_department.get(department_key, 0) + 1
            if status_key not in {"verified", "closed"}:
                item = row_to_dict(finding)
                item["station_name"] = station_names.get(finding.station_code)
                open_findings.append(item)
                target = parse_date_value(finding.target_date)
                if target and target < today:
                    overdue_findings.append({**item, "days_overdue": (today - target).days})

        return {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "as_of": today.isoformat(),
            "overview": {
                "stations": len(stations),
                "units": len(units),
                "active_units": len(active_units),
                "earnings_total": sum(to_money(row.amount) for row in earnings),
                "works": len(works_seen),
                "open_works": pending_work_count,
                 "critical_alerts": bucket_counts["overdue"] + bucket_counts["needs_review"] + len(overdue_findings),
            },
            "stations": {
                "total": len(stations),
                "with_units": len(station_with_units),
                "without_units": max(len(station_codes - station_with_units), 0),
                "with_works": len(station_with_works),
                "with_earnings": len(station_with_earnings),
                "by_category": [{"label": key, "value": value} for key, value in sorted(by_station_category.items())],
                "by_division": [{"label": key, "value": value} for key, value in sorted(by_division.items())],
                "by_section": [{"label": key, "value": value} for key, value in sorted(by_section.items())],
            },
            "units": {
                "total": len(units),
                "active": len(active_units),
                "inactive": inactive_units,
                "missing_license_fee": sum(1 for unit, *_ in active_units if not to_money(unit.license_fee)),
                "by_status": [{"label": key, "value": value} for key, value in sorted(by_status.items())],
                "by_type": [{"label": key, "value": value} for key, value in sorted(by_type.items())],
                "by_category": [{"label": key, "value": value} for key, value in sorted(by_unit_category.items())],
            },
            "earnings": {
                "total_receipts": len(earnings),
                "total_amount": sum(to_money(row.amount) for row in earnings),
                "license_fee_receipts": len(license_earnings),
                "license_fee_collected": sum(to_money(row.amount) for row in license_earnings),
                "pending_receipts": len(pending_receipts),
                "pending_receipt_amount": sum(to_money(row.amount) for row in pending_receipts),
                "current_month_license_collected": sum(to_money(row.amount) for row in current_month_license),
                "last_3_month_license_collected": sum(to_money(row.amount) for row in recent_license),
                "by_month": [{"label": key, "value": value} for key, value in sorted(earnings_by_month.items())[-12:]],
                "by_head": [{"label": key, "value": value} for key, value in sorted(earnings_by_head.items(), key=lambda item: item[1], reverse=True)[:10]],
            },
            "works": {
                "total": len(works_seen),
                "completed": completed_work_count,
                "pending": pending_work_count,
                "by_status": [{"label": key, "value": value} for key, value in sorted(works_by_status.items(), key=lambda item: item[1], reverse=True)],
                "by_scope": [{"label": key, "value": value} for key, value in sorted(works_by_scope.items(), key=lambda item: item[1], reverse=True)],
                "by_section": [{"label": key, "value": value} for key, value in sorted(works_by_section.items(), key=lambda item: item[1], reverse=True)[:12]],
                "by_station": [{"label": key, "value": value} for key, value in sorted(works_by_station.items(), key=lambda item: item[1], reverse=True)[:12]],
                "by_sr_den": [{"label": key, "value": value} for key, value in sorted(works_by_sr_den.items(), key=lambda item: item[1], reverse=True)],
                "by_cmi": [{"label": key, "value": value} for key, value in sorted(works_by_cmi.items(), key=lambda item: item[1], reverse=True)],
                "by_allocation": [{"label": key, "value": value} for key, value in sorted(works_by_allocation.items(), key=lambda item: item[1], reverse=True)],
                "by_year": [{"label": key, "value": value} for key, value in sorted(works_by_year.items(), key=lambda item: item[0])],
                "delayed": len(delay_alerts),
                "contradictions": len(work_contradictions),
                "delay_alerts": sorted(delay_alerts, key=lambda row: row["days_overdue"], reverse=True),
                "contradiction_rows": work_contradictions,
            },
            "data_quality": {
                "units_missing_station": sum(1 for unit, *_ in units if unit.station_code and unit.station_code not in station_codes),
                "earnings_missing_unit": sum(1 for row in earnings if row.unit_no and row.unit_no not in unit_codes),
                "earnings_missing_station": sum(1 for row in earnings if row.station_code and row.station_code not in station_codes),
                "works_unmatched_station": sum(1 for _, link, *_ in works if link and link.scope_type == "Station" and link.station_code and link.station_code not in station_codes),
                "units_missing_license_fee": sum(1 for unit, *_ in units if not to_money(unit.license_fee)),
            },
            "license_fee_alerts": {
                **bucket_counts,
                "estimated_overdue_amount": estimated_overdue_amount,
                "rows": alerts[:300],
            },
            "commercial_contracts": commercial_reports,
            "inspections": {
                "total": len(inspections),
                "by_status": [{"label": key, "value": value} for key, value in sorted(inspection_by_status.items(), key=lambda item: item[1], reverse=True)],
                "findings_open": len(open_findings),
                "findings_overdue": len(overdue_findings),
                "by_status_findings": [{"label": key, "value": value} for key, value in sorted(finding_by_status.items(), key=lambda item: item[1], reverse=True)],
                "by_severity": [{"label": key, "value": value} for key, value in sorted(finding_by_severity.items(), key=lambda item: item[1], reverse=True)],
                "by_department": [{"label": key, "value": value} for key, value in sorted(finding_by_department.items(), key=lambda item: item[1], reverse=True)],
                "overdue_findings": sorted(overdue_findings, key=lambda row: row["days_overdue"], reverse=True)[:300],
                "open_findings": open_findings[:300],
            },
        }
    finally:
        session.close()
