"""Import CMI, DEN and Sr DEN from the authoritative Stations worksheet.

Usage from the backend directory:
    python scripts/sync_station_assignments.py

The operation is deliberately limited to officer-assignment columns so it
cannot overwrite station master or passenger-amenity data accidentally.
"""

from __future__ import annotations

import argparse
import io
import re
import sys
from pathlib import Path
from urllib.request import urlopen

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from database import SessionLocal  # noqa: E402
from models import Station  # noqa: E402


SOURCE_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1UdRgQQPEkak1fUTuVH7jIn5R4sE3szAhM4VZJOdFIOU/export?format=xlsx"
)


def key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def value(row: tuple[object, ...], indexes: dict[str, int], *names: str) -> str | None:
    for name in names:
        index = indexes.get(key(name))
        if index is not None and index < len(row):
            text = str(row[index] or "").strip()
            return text or None
    return None


def read_source() -> list[dict[str, str | None]]:
    with urlopen(SOURCE_URL, timeout=60) as response:
        workbook = load_workbook(io.BytesIO(response.read()), read_only=True, data_only=True)
    if "Stations" not in workbook.sheetnames:
        raise RuntimeError("The source workbook has no Stations worksheet")
    rows = list(workbook["Stations"].iter_rows(values_only=True))
    if not rows:
        return []
    indexes = {key(header): index for index, header in enumerate(rows[0])}
    missing = [name for name in ("station code", "cmi", "den", "sr den") if key(name) not in indexes]
    if missing:
        raise RuntimeError(f"Stations worksheet is missing columns: {', '.join(missing)}")
    result: list[dict[str, str | None]] = []
    for row in rows[1:]:
        station_code = value(row, indexes, "station code")
        if station_code:
            result.append({
                "station_code": station_code.upper(),
                "cmi": value(row, indexes, "cmi"),
                "den": value(row, indexes, "den"),
                "sr_den": value(row, indexes, "sr den", "sr.den"),
            })
    return result


def sync(*, dry_run: bool = False) -> tuple[int, int]:
    source_rows = read_source()
    session = SessionLocal()
    updated = created = 0
    try:
        for item in source_rows:
            station = session.get(Station, item["station_code"])
            if station is None:
                # Do not create incomplete station masters from an assignment-only import.
                continue
            station.cmi = item["cmi"]
            station.den = item["den"]
            station.sr_den = item["sr_den"]
            updated += 1
        if dry_run:
            session.rollback()
        else:
            session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()
    return updated, created


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Fetch and validate without committing")
    args = parser.parse_args()
    updated, created = sync(dry_run=args.dry_run)
    mode = "would update" if args.dry_run else "updated"
    print(f"{mode} {updated} existing stations; skipped missing station masters; created {created}")


if __name__ == "__main__":
    main()
